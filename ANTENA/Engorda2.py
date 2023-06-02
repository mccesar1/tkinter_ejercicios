
import binascii
import os
from sqlite3 import Row
import threading
import time
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import pandas as pd 
import openpyxl

import socket

class Principal:
  try:
     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     server_address = ('192.168.0.231', 7086)
    #server_address = ('192.168.1.200', 100)
     sock.connect(server_address)
     estado=1
  except:
     estado=0

  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus')
        #self.ws.geometry('1000x600')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='#044275')
        self.widgets()
        s = ttk.Style()
        s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, background='#ffffff') 
        s.configure('TNotebook', background='#ffffff')
        s.configure('TFrame', background='#ffffff')
        s.configure('TLabel', background='#ffffff')
        s.configure('TButton', background='#ffffff')

####funcion para conectar la antena####################################################################################################################   
 
  def sockets(self,estado):
    # try:
    #     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    #     server_address = ('192.168.0.238', 7086)
    #server_address = ('192.168.1.200', 100)
       # sock.connect(server_address)
       if estado==1:
        self.EestadoAntena2.insert(0, 'Conectada')

       else:
        self.EestadoAntena2.insert(0, 'Error')

  def select_usb(self):
        usb = os.system('fsutil fsinfo drives')
        usb = usb.split('\n')
        listaUSB = [ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:']

        print ("cesar ",usb)
        return listaUSB

##export xls##################################################################################################################################
  def guardar_excel(self): 
        wb = openpyxl.Workbook()    #crea el libro de excel
        sheet = wb.active  #selecciona la hoja activa
        sheet.title = 'Tags' #se le asigna un nombre a la hoja
        sheet['A1'] = 'Tags' #se le asigna un titulo a la columna A
        sheet['B1'] = 'Fecha' #se le asigna un titulo a la columna B
        sheet['C1'] = 'Hora' #se le asigna un titulo a la columna C
        
        #se le asigna un valor a cada celda
        for i in range(2, len(self.btn.get("1.0", "end-1c").splitlines())+2): 
            sheet.cell(row=i, column=1).value = self.btn.get("1.0", "end-1c").splitlines()[i-2] # i-2 para que no se repitan los datos
            sheet.cell(row=i, column=2).value = time.strftime("%d/%m/%Y")
            sheet.cell(row=i, column=3).value = time.strftime("%H:%M:%S")
            
            fecha = time.strftime("%d-%m-%Y")
            hora = time.strftime("%Hhrs%Mmin")
            extension = ".xlsx"

        selection = self.cmb_usb.get()
        if selection == '':
            messagebox.showinfo(message="Seleccione una unidad de disco", title="Selección")
       
        #wb.save(f'tags{hora}{fecha}{extension}') #se guarda el archivo en la carpeta actual
        else:
            try:
                 wb.save(selection+'\Tags-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
                 messagebox.showinfo("Excel", "Se ha guardado el archivo excel")
            except:
                messagebox.showinfo(message="No se encontro la unidad usb", title="Error")
        

####funccion para iniciar la lectura#####################################################################################     
  def IniciarLectura(self):
       
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        #Read = bytes([0x0A, 0x00, 0x02, 0x28, 0x34])
        #    = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(Read)
        hilo1=threading.Thread(target=self.GetData).start()
  
  def GetData(self):
        lista = []

        while True:
         data = self.sock.recv(4096)  
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             #print ("Card Scanned. Tag ID:", data)
             if data not in lista:
              lista.append(data)
              self.btn.insert(tk.END, data+"\n") #inserta los datos en el textbox

####funccion para detener la lectura##############################################################################################################################

  def DetenerLectura(self):    
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
        
       # self.grid3.insert(tk.END, self.lista, self.lista.index(i+1))
####funcion para salir##################################################################################################################################
  def salir(self):
        self.ws.destroy()


#####funcion para crear widgets##################################################################################################################################
  def widgets(self):

    #FRAME DE LA IZQUIERDA###########################################################
    left_frame = Frame(ws, bd=3, relief=SOLID, padx=210, pady=340, background='#044275', highlightbackground="white", highlightthickness=3)
    left_frame.place(x=10, y=10)
    self.btn = tk.Text(left_frame, width=10, height=20, font=('Arial', 32), border=3, relief=tk.GROOVE, background='#ffffff')
    self.btn.place(x=-213, y=-342, width=620, height=900)

  
    ####FRAME DE LA DERECHA#########################################################################
    right_frame = Frame(ws, bd=3,  background='#044275', relief=SOLID, padx=320, pady=340)
    right_frame.place(x=650, y=10)
    re_pw = Label(right_frame, font=('fsdfds', 20), background='#044275')
    re_pw.grid(row=6, column=1, pady=10, padx=20)
    

    self.cmb_usb=ttk.Combobox(right_frame, width=10, justify='center', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:', 'J:'], state='readonly')
    self.cmb_usb.grid(row=5, column=1, ipady=9, ipadx=20)
    EestadoAntena = Entry(right_frame, font=('ARIAL', 28), background='#ffffff', justify='center')
    EestadoAntena.place(x=-170, y=150, width=440, height=60)
    EestadoAntena.insert(0, 'Estado de la antena')
    separador = ttk.Separator(right_frame, orient=HORIZONTAL)
    separador.place(x=-260, y=225, width=530, height=3)
    self.EestadoAntena2 = Entry(right_frame, font=('ARIAL', 24), background='#ffffff', justify='center')
    self.EestadoAntena2.place(x=-260, y=240, width=530, height=80)
    self.EestadoAntena2.insert(0, 'Conectada')
    login_btn = Label(left_frame, width=15, text='', font=('Times', 14), command=None, background='#ffffff')
    login_btn.grid(row=20, column=1, pady=10, padx=20)
    Euid = Entry(right_frame, font=('Times', 24), background='#b8b8b8')
    Euid.place(x=-260, y=150, width=80, height=60)

  

    Bxls = Button(right_frame, text='XLS', font=('ARIAL', 28), background='#b8b8b8', command=self.guardar_excel)
    Bxls.place(x=-260, y=-180, width=200, height=80) 
    Bleer=Button(right_frame, text="Leer", font=('ARIAL', 28), background='#b8b8b8', command=self.IniciarLectura)
    Bleer.place(x=-260, y=-280, width=200, height=80)

    Bsalir = Button(right_frame, text='SALIR', font=('ARIAL', 28), background='#b8b8b8', command=self.salir)
    Bsalir.place(x=90, y=-180, width=200, height=80)
    Bparar=Button(right_frame, text="Parar", font=('ARIAL', 28), background='#b8b8b8', command=self.DetenerLectura)
    Bparar.place(x=90, y=-280, width=200, height=80)

    


if __name__ == '__main__':
    ws = Tk()
    application = Principal(ws)
    ws.mainloop()