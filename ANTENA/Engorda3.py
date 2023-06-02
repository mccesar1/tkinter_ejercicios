import binascii
import os
import threading
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import openpyxl
import time
import socket

class Principal:
  try:
     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     server_address = ('192.168.0.231', 7086)
    #server_address = ('192.168.1.200', 100)
     sock.connect(server_address)
     estado=1
  except:
     estado=0

  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus')
        #self.ws.geometry('1000x600')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='#003C78')
        self.widgets()
        s = ttk.Style()
        s.map('TCombobox', fieldbackground=[('readonly','black')])
        s.map('TCombobox', selectbackground=[('readonly', 'black')])
        s.map('TCombobox', selectforeground=[('readonly', 'white')])
        s.map('TCombobox', foreground=[('readonly', 'white')])
        s.map('TCombobox', background=[('readonly', 'black')])
        s.map('TCombobox', arrowcolor=[('readonly', 'white')])
        s.map('TCombobox', selectborderwidth=[('readonly', '0')])
        s.map('TCombobox', bordercolor=[('readonly', 'black')])
        s.map('TCombobox', borderwidth=[('readonly', '0')])
        s.map('TCombobox', relief=[('readonly', 'flat')])
        #s.configure("TCombobox", fieldbackground= "black", background= "#003C78", foreground= "black", selectbackground= "black", selectforeground= "white", arrowcolor= "black", bordercolor= "black", lightcolor= "black", darkcolor= "black", troughcolor= "black", focuscolor= "black", selectcolor= "black")
        s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, fieldbackground= "orange", background='#000000', foreground='#000000', borderwidth=2)
        s.configure('TNotebook', background='#003C78')
        s.configure('TFrame', background='#003C78')
        s.configure('TLabel', background='#003C78')
        s.configure('TButton', background='#000000')

####funcion para conectar la antena####################################################################################################################   
  def sockets(self,estado):
       if estado==1:
        self.EestadoAntena2.insert(0, 'Conectada')
       else:
        self.EestadoAntena2.insert(0, 'Error')

####funccion para iniciar la lectura#####################################################################################     
  def IniciarLectura(self):
       
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        self.sock.send(Read)
        hilo1=threading.Thread(target=self.GetData).start()
  
  def GetData(self):
        self.lista = []
        leidas = 5
        while True:
         data = self.sock.recv(4096)  
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             #print ("Card Scanned. Tag ID:", data)
             if data not in self.lista:
              self.lista.append(data)
              #print ("guardadas",lista)
              leidas = (len(self.lista))
              self.Eleidas.delete(0, END)  #borra el contenido del textbox de 'leidas' para que no se repitan los datos
              self.Eleidas.insert(tk.END, leidas) #inserta el numero de datos leidos en el textbox
              self.btn.insert(tk.END, data+"\n")
            for leidas in self.lista:
                self.Earete_entry.delete(0, END)
                self.Earete_entry.insert(tk.END, leidas)
                self.Euid2.delete(0, END)
                self.Euid2.insert(tk.END, data)

####funccion para detener la lectura##############################################################################################################################
  def DetenerLectura(self):    
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)

####funcion para salir##################################################################################################################################
  def salir(self):
        self.ws.destroy()

  def select_usb(self):
        usb = os.system('fsutil fsinfo drives')
        usb = usb.split('\n')
        listaUSB = [ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:']
        print ("cesar ",usb)
        return listaUSB

##export xls##################################################################################################################################
  def guardar_excel(self): 
        wb = openpyxl.Workbook()    #crea el libro de excel
        sheet = wb.active  #selecciona la hoja activa
        sheet.title = 'Tags' #se le asigna un nombre a la hoja
        sheet['A1'] = 'Tags' #se le asigna un titulo a la columna A
        sheet['B1'] = 'Fecha' #se le asigna un titulo a la columna B
        sheet['C1'] = 'Hora' #se le asigna un titulo a la columna C
        #se le asigna un valor a cada celda
        for i in range(2, len(self.btn.get("1.0", "end-1c").splitlines())+2): 
            sheet.cell(row=i, column=1).value = self.btn.get("1.0", "end-1c").splitlines()[i-2] # i-2 para que no se repitan los datos
            sheet.cell(row=i, column=2).value = time.strftime("%d/%m/%Y")
            sheet.cell(row=i, column=3).value = time.strftime("%H:%M:%S")
            fecha = time.strftime("%d-%m-%Y")
            hora = time.strftime("%Hhrs%Mmin")
        selection = self.cmb_usb.get()
        if selection == '':
            messagebox.showinfo(message="Seleccione una unidad de disco", title="Selección")
        else:
            try:
                 wb.save(selection+'\Tags-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
                 messagebox.showinfo("Excel", "Se ha guardado el archivo excel")
            except:
                messagebox.showinfo(message="No se encontro la unidad usb", title="Error")
#####funcion para crear widgets##################################################################################################################################
  def widgets(self):
#####FRAME DE LA IZQUIERDA###########################################################
    left_frame = Frame(ws, bd=3, relief=SOLID, padx=210, pady=340, background='#003C78', highlightbackground="white", highlightthickness=3)
    left_frame.place(x=10, y=10)
    Luid=Label(left_frame, text="UID", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    Luid.place(x=-160, y=-280)
    Larete_uhf=Label(left_frame, text="Arete UHF", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    Larete_uhf.place(x=-10, y=-280)
    self.Earete_entry = Entry(left_frame, font=('ARIAL', 24), background='#ccd7e0')
    self.Earete_entry.place(x=-10, y=-200, width=380, height=60)
    Euid = Entry(left_frame, font=('Times', 24), background='#ccd7e0')
    Euid.place(x=-160, y=-200, width=130, height=60)
    self.Euid2 = Entry(left_frame, font=('ARIAL', 24), background='#070807', foreground='#00ff80')
    self.Euid2.place(x=-10, y=-110, width=380, height=60)
    self.Earete_uhf2 = Entry(left_frame, font=('Times', 24), background='#070807', foreground='#00ff80')
    self.Earete_uhf2.place(x=-160, y=-110, width=130, height=60)
    Euid = Entry(left_frame, font=('Arial', 24), background='#b8b8b8', justify='center')
    Euid.place(x=-160, y=150, width=80, height=60)
    Euid.insert(tk.END, "7")
    EestadoAntena = Entry(left_frame, font=('ARIAL', 28), background='#ffffff', justify='center')
    EestadoAntena.place(x=-70, y=150, width=440, height=60)
    EestadoAntena.insert(0, 'Estado de la antena')
    separador = ttk.Separator(left_frame, orient=HORIZONTAL)
    separador.place(x=-160, y=225, width=530, height=3)
    self.EestadoAntena2 = Entry(left_frame, font=('ARIAL', 30), background='#070807', justify='center', foreground='#1fb1d1')
    self.EestadoAntena2.place(x=-160, y=240, width=530, height=80)
    self.EestadoAntena2.insert(0, 'Conectada')
    login_btn = Label(left_frame, width=15, text='', font=('Times', 14), command=None, background='#003C78')
    login_btn.grid(row=20, column=1, pady=10, padx=20)
########FRAME DE LA DERECHA##############################################################################################
    right_frame = Frame(ws, bd=3,  background='#003C78', relief=SOLID, padx=320, pady=340)
    right_frame.place(x=650, y=10)
    re_pw = Label(right_frame, font=('', 14))
    re_pw.grid(row=6, column=1, pady=10, padx=20)
    
    tabControl = ttk.Notebook(right_frame)
    tab1 = ttk.Frame(tabControl, width=100, height=20)
    tab4 = ttk.Frame(tabControl, width=100, height=20)

    tabControl.add(tab1, text ='Captura')
    tabControl.add(tab4, text ='Leidas')
    tabControl.place(x=-320, y=-340, width=687, height=590)

    Bxls = Button(right_frame, text='XLS', font=('ARIAL', 24), background='#DCE1D8', command=self.guardar_excel)
    Bxls.place(x=-290, y=280, width=200, height=80)
    Bsalir = Button(right_frame, text='Salir', font=('ARIAL', 24), background='#DCE1D8', command=self.salir)
    Bsalir.place(x=90, y=280, width=200, height=80)
####tab1###########################################################################################
    Larete_uhf=Label(tab1, text="Contador", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    Larete_uhf.place(x=30, y=30)
    self.Eleidas=Entry(tab1, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    self.Eleidas.place(x=30, y=80, width=200, height=60)
    self.Eleidas.insert(0, '0')
    Larete_uhf=Label(tab1, text="USB", font=('ARIAL', 28), background='#003C78', foreground='#ffffff')
    Larete_uhf.place(x=30, y=250)
    self.cmb_usb=ttk.Combobox(tab1, width=10, font=('ARIAL', 28), justify='left', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:'], state='readonly')
    self.cmb_usb.place(x=30, y=300, width=200, height=60)

    Bleer=Button(tab1, text="Leer", font=('ARIAL', 28), background='#DCE1D8', command=self.IniciarLectura)
    Bleer.place(x=410, y=60, width=200, height=120)
    Bparar=Button(tab1, text="Parar", font=('ARIAL', 28), background='#DCE1D8', command=self.DetenerLectura)
    Bparar.place(x=410, y=220, width=200, height=120)

####tab4###########################################################################################
    self.btn = tk.Text(tab4, width=10, height=20, font=('Arial', 32), border=3, relief=tk.GROOVE, background='#070807', foreground='#00ff80')
    self.btn.place(x=0, y=0, width=720, height=900)

if __name__ == '__main__':
    ws = Tk()
    application = Principal(ws)
    ws.mainloop()