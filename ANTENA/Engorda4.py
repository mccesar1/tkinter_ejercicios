import binascii
import os
import threading
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import openpyxl
import time
import socket
import os
import sqlite3

class Principal:
  db_name = 'engorda.db'
  try:
     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     server_address = ('192.168.0.231', 7086)
    #server_address = ('192.168.1.200', 100)
     sock.connect(server_address)
     estado=1
  except:
     estado=0

  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus')
        #self.ws.geometry('1000x600')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='#003C78')
        self.widgets()
        self.VerDatos()
        s = ttk.Style()
        #s.theme_use('clam')
        s.theme_use('alt')
        self.respaldo=0
        s.map('TCombobox', fieldbackground=[('readonly','black')])
        s.map('TCombobox', selectbackground=[('readonly', 'black')])
        s.map('TCombobox', selectforeground=[('readonly', 'white')])
        s.map('TCombobox', foreground=[('readonly', 'white')])
        s.map('TCombobox', background=[('readonly', 'black')])
        s.map('TCombobox', arrowcolor=[('readonly', 'white')])
        s.map('TCombobox', selectborderwidth=[('readonly', '0')])
        s.map('TCombobox', bordercolor=[('readonly', 'black')])
        s.map('TCombobox', borderwidth=[('readonly', '0')])
        s.map('TCombobox', relief=[('readonly', 'flat')])
        s.configure("TCombobox", fieldbackground= "black", background= "black", foreground= "black", selectbackground= "black", selectforeground= "black", arrowcolor= "black", bordercolor= "black", lightcolor= "black", darkcolor= "black", troughcolor= "black", focuscolor= "black", selectcolor= "black")
        s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, fieldbackground= "orange", background='black', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TNotebook', background='#003C78', borderwidth=0, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TFrame', background='#003C78')
        s.configure('TLabel', background='#003C78')
        s.configure('TButton', background='#000000', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='white')
        
        s.configure("Treeview.Heading", font=(None, 30))
        s.configure("Treeview.Heading", background="#993C78")
        s.configure("Treeview", font=(None, 30))
        s.configure("Treeview", rowheight=40)
        ttk.Style().configure("Treeview", background="black", 
                fieldbackground="black", foreground="white")
        #s.configure('TFrame', background='#black')
        for line in self.grid3.get_children():
         for value in self.grid3.item(line)['values']:
          self.lista2.append(str(value))
          #print(self.lista2)

####funcion para conectar la antena####################################################################################################################   
  def sockets(self,estado):
       if estado==1:
        self.EestadoAntena2.insert(0, 'Conectada')
       else:
        self.EestadoAntena2.insert(0, 'Error')

####funccion para iniciar la lectura#####################################################################################     
  def IniciarLectura(self):      
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        self.sock.send(Read)
        hilo1=threading.Thread(target=self.GetData).start()
#######Funcion para crear la base de datos #############################################################################################################
  def create_database(self):
        self.conn = sqlite3.connect('engorda.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS tags (id INTEGER, tag TEXT)')
        self.conn.commit()
#### Funcion para hacer la conexio###########################################################################################################################
  def conexion(self, query, parameters):
     try:
      with sqlite3.connect(self.db_name) as conn:
        cursor = conn.cursor()
        result = cursor.execute(query, (parameters,))
        conn.commit()
        return result
     except Exception as e:
        messagebox.showerror("ERROR", "Error al conectar con la base de datos")
        self.create_database()
        #print("Error de conexion: ", e)
        return None
######funcion para guardar los datos en la base de datos###############################################################################################        
  def deleteRegister(self):
    # query = 'DELETE FROM tags '
    # self.conexion(query, '')
    # messagebox.showinfo('AGROPLUS', 'Los datos han sido eliminados de la base de datos')
    # #self.VerDatos()
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'DELETE FROM tags'
      cursor.execute(query)
      conn.commit()
      cursor.close()
      
  def GuardarEnBD(self):
        self.deleteRegister()
        for item in self.grid3.get_children():
          # query = 'INSERT INTO tags VALUES (NULL, ?)'
            query = 'INSERT INTO tags VALUES (NULL, ?)'
            parameters = (self.grid3.item(item)['values'][0])
            #parameters2 = (self.grid3.item(item)['values'][1])
            #print (self.grid3.item(item)['values'][1])
            self.conexion(query, parameters)
        messagebox.showinfo('AGROPLUS', 'Los datos han sido guardados en la base de datos')
        self.VerDatos
        self.respaldo=1

  def VerDatos(self):
    self.lista2 = []
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM tags'
      result = cursor.execute(query)
      conn.commit()
      for row in result:
        #  self.lista2.append(row[1])
        
         self.grid3.insert('',0,values=row[1])
         #self.grid3.insert('',0,values=row[1], text=row[0])
      cursor.close()


#####Funcion para obtener los datos de la antena###########################################################################################
  def GetData(self):
        self.lista = []        
        leidas = 5
        i=1
        while True:
         data = self.sock.recv(4096)  
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             if data not in self.lista :
              self.lista.append(data)
              leidas = (len(self.lista))
              
              self.Eleidas.delete(0, END)  #borra el contenido del textbox de 'leidas' para que no se repitan los datos
              self.Eleidas.insert(tk.END, leidas) #inserta el numero de datos leidos en el textbox
              #self.btn.insert(tk.END, data+"\n")
              if data not in self.lista2:
              
               self.grid3.insert('',0,text=""+str(i),values=data)
               i=i+1
              # self.grid3.insert('', 'end', text=""+str(i), values=data)
              # i=i+1
            #   self.btn.insert(tk.END, data+"\n")

              for column in self.grid3['columns']:
                self.grid3.heading(column, text=column)               
            #   for column in self.grid3['columns2']:
            #     self.grid3.heading(column, text=column)
            #   for data in self.lista:
            #    self.grid3.insert('', 'end', values=data)
            #    #self.grid3.insert()
               
            for leidas in self.lista:
                self.Earete_entry.delete(0, END)
                self.Earete_entry.insert(tk.END, leidas)
                self.Euid2.delete(0, END)
                #self.Earete.delete(0, END)
                
                self.Euid2.insert(tk.END, data) 
                #self.Earete.insert(tk.END, data)
                 #self.Euid2.delete(0, END)
            

####funccion para detener la lectura##############################################################################################################################
  def DetenerLectura(self):  
        
        self.Euid2.delete(0, END)
        self.Euid2.insert(tk.END, "") 
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
       
####funcion para salir##################################################################################################################################
  def salir(self):
    if self.respaldo==1:

        self.DetenerLectura()
        self.sock.close()
        time.sleep(0.1) 
        self.ws.destroy()
    else:
        messagebox.showerror("ERROR", "No se han guardado los datos en la base de datos")
   

  def select_usb(self):
        usb = os.system('fsutil fsinfo drives')
        usb = usb.split('\n')
        listaUSB = [ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:']
        print ("cesar ",usb)
        return listaUSB

##export xls##################################################################################################################################
  def treeViewToExcel(self):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws['A1'] = 'Tags' #se le asigna un titulo a la columna A
    ws['B1'] = 'Fecha' #se le asigna un titulo a la columna B
    ws['C1'] = 'Hora' #se le asigna un titulo a la columna C
    
    ws.append(self.grid3['columns'])
    selection = self.cmb_usb.get()
    fecha = time.strftime("%d-%m-%Y")
    hora = time.strftime("%Hhrs%Mmin")

    for row in self.grid3.get_children():
      ws.append(self.grid3.item(row)['values'])
      
      # ws.cell(row=i, column=2).value = time.strftime("%d/%m/%Y")
      # ws.cell(row=i, column=3).value = time.strftime("%H:%M:%S")
    
    wb.save(selection+'Datos'+fecha+'-'+hora+'.xlsx' )
    messagebox.showinfo('AGROPLUS', 'Los datos han sido exportados a Excel')
    #self.VerDatos()

  

  def guardar_excel(self): 
        wb = openpyxl.Workbook()    #crea el libro de excel
        sheet = wb.active  #selecciona la hoja activa
        sheet.title = 'Tags' #se le asigna un nombre a la hoja
        sheet['A1'] = 'Tags' #se le asigna un titulo a la columna A
        sheet['B1'] = 'Fecha' #se le asigna un titulo a la columna B
        sheet['C1'] = 'Hora' #se le asigna un titulo a la columna C
        #se le asigna un valor a cada celda
        for i in range(2, len(self.btn.get("1.0", "end-1c").splitlines())+2): 
            sheet.cell(row=i, column=1).value = self.btn.get("1.0", "end-1c").splitlines()[i-2] # i-2 para que no se repitan los datos
            sheet.cell(row=i, column=2).value = time.strftime("%d/%m/%Y")
            sheet.cell(row=i, column=3).value = time.strftime("%H:%M:%S")  
            fecha = time.strftime("%d-%m-%Y")
            hora = time.strftime("%Hhrs%Mmin")
        selection = self.cmb_usb.get()
        if selection == '':
            messagebox.showinfo(message="Seleccione una unidad de disco", title="Selección")
        else:
            try:
                 wb.save(selection+'\Tags-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
                 messagebox.showinfo("Excel", "Se ha guardado el archivo excel")
            except:
                messagebox.showinfo(message="No se encontro la unidad usb", title="Error")

  def DeleteAll(self):
        self.btn.delete('1.0', END)
        #self.grid3.delete(*self.grid3.get_children())
        self.lista = []
        self.Eleidas.delete(0, END)
        self.Eleidas.insert(tk.END, 0)
        self.Earete_entry.delete(0, END)
        self.Euid2.delete(0, END)
        #self.deleteRegister()
        #self.lista2 = []
        #self.lista1 = []

  def DropTable(self):
        self.conexion("DROP TABLE tags")
        messagebox.showinfo("AGROPLUS", "La tabla ha sido eliminada")
        self.VerDatos()   

  def ConfirmEXIT(self):
         if messagebox.askokcancel("Salir", "RECUERDE RESPALDAR LOS DATOS ANTES DE SALIR"):
            self.salir()                 
#####funcion para crear widgets##################################################################################################################################
  def widgets(self):
#####FRAME DE LA IZQUIERDA###########################################################
    left_frame = Frame(ws, bd=3, relief=SOLID, padx=210, pady=340, background='#003C78', highlightbackground="white", highlightthickness=3)
    left_frame.place(x=10, y=10)
    Luid=Label(left_frame, text="UID", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    Luid.place(x=-160, y=-280)
    Larete_uhf=Label(left_frame, text="Arete UHF", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    Larete_uhf.place(x=-10, y=-280)
    self.Earete_entry = Entry(left_frame, font=('ARIAL', 34), background='#ccd7e0')
    self.Earete_entry.place(x=-10, y=-200, width=380, height=60)
    Euid = Entry(left_frame, font=('Times', 24), background='#ccd7e0')
    Euid.place(x=-160, y=-200, width=130, height=60)
    self.Euid2 = Entry(left_frame, font=('ARIAL', 34), background='#070807', foreground='#00ff80')
    self.Euid2.place(x=-10, y=-110, width=380, height=60)
    self.Earete_uhf2 = Entry(left_frame, font=('Times', 24), background='#070807', foreground='#00ff80')
    self.Earete_uhf2.place(x=-160, y=-110, width=130, height=60)
    Euid = Entry(left_frame, font=('Arial', 24), background='#b8b8b8', justify='center')
    Euid.place(x=-160, y=150, width=80, height=60)
    Euid.insert(tk.END, "7")
    EestadoAntena = Entry(left_frame, font=('ARIAL', 28), background='#ffffff', justify='center')
    EestadoAntena.place(x=-70, y=150, width=440, height=60)
    EestadoAntena.insert(0, 'Estado de la antena')
    separador = ttk.Separator(left_frame, orient=HORIZONTAL)
    separador.place(x=-160, y=225, width=530, height=3)
    self.EestadoAntena2 = Entry(left_frame, font=('ARIAL', 30), background='#070807', justify='center', foreground='#1fb1d1')
    self.EestadoAntena2.place(x=-160, y=240, width=530, height=80)
    self.EestadoAntena2.insert(0, 'Conectada')
    login_btn = Label(left_frame, width=15, text='', font=('Times', 14), command=None, background='#003C78')
    login_btn.grid(row=20, column=1, pady=10, padx=20)

########FRAME DE LA DERECHA##############################################################################################
    right_frame = Frame(ws, bd=3,  background='#003C78', relief=SOLID, padx=320, pady=340)
    right_frame.place(x=650, y=10)
    re_pw = Label(right_frame, font=('', 14))
    re_pw.grid(row=6, column=1, pady=10, padx=20)
    
    tabControl = ttk.Notebook(right_frame)
    tab1 = ttk.Frame(tabControl, width=100, height=20)
    
    tab5 = ttk.Frame(tabControl, width=100, height=20)
    tab6 = ttk.Frame(tabControl, width=100, height=20)

    tabControl.add(tab1, text ='Lectura')
    
    tabControl.add(tab5, text ='Leidas')
    tabControl.add(tab6, text ='Captura')
    tabControl.place(x=-320, y=-340, width=687, height=590)

    Bxls = Button(right_frame, text='XLS', font=('ARIAL', 24), background='#DCE1D8', command=self.treeViewToExcel, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bxls.place(x=-290, y=280, width=200, height=80)
    Bsalir = Button(right_frame, text='Salir', font=('ARIAL', 24), background='#DCE1D8', command=self.salir, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bsalir.place(x=90, y=280, width=200, height=80)

####tab1###########################################################################################
    Larete_uhf=Label(tab1, text="Contador", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    Larete_uhf.place(x=30, y=30)
    self.Eleidas=Entry(tab1, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    self.Eleidas.place(x=30, y=80, width=200, height=60)
    self.Eleidas.insert(0, '0')
    Larete_uhf=Label(tab1, text="USB", font=('ARIAL', 28), background='#003C78', foreground='#ffffff')
    Larete_uhf.place(x=30, y=310)
    self.cmb_usb=ttk.Combobox(tab1, width=10, font=('ARIAL', 28), justify='left', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:'], state='readonly')
    self.cmb_usb.place(x=30, y=380, width=200, height=60)
    self.cmb_usb.current(1)
    Bleer=Button(tab1, text="Leer", font=('ARIAL', 28), background='#DCE1D8', command=self.IniciarLectura, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bleer.place(x=410, y=50, width=200, height=120)
    Bparar=Button(tab1, text="Parar", font=('ARIAL', 28), background='#DCE1D8', command=self.DetenerLectura, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bparar.place(x=410, y=210, width=200, height=120)
    Brespaldo=Button(tab1, text="Respaldo", font=('ARIAL', 28), background='#DCE1D8', command=self.GuardarEnBD, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Brespaldo.place(x=410, y=370, width=200, height=120)
    Bzero=Button(tab1, text="Zero", font=('ARIAL', 28), background='#DCE1D8', command=self.DeleteAll, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6, relief=RIDGE) 
    Bzero.place(x=30, y=200, width=200, height=80)

###tab6###########################################################################################
    self.btn = tk.Text(tab6, width=10, height=20, font=('Arial', 32), border=3, relief=tk.GROOVE, background='#070807', foreground='#00ff80')
    self.btn.place(x=0, y=0, width=720, height=900)
    self.Lpeso = Label(tab6, text="Peso", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    self.Lpeso.place(x=30, y=30)
    self.Epeso=Entry(tab6, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    self.Epeso.place(x=30, y=80, width=200, height=60)
    self.Ledad = Label(tab6, text="Edad", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    self.Ledad.place(x=30, y=150)
    self.Eedad=Entry(tab6, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    self.Eedad.place(x=30, y=200, width=200, height=60)
    self.Lsexo = Label(tab6, text="Sexo", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    self.Lsexo.place(x=30, y=270)
    self.cmb_sexo=ttk.Combobox(tab6, width=10, font=('ARIAL', 28), justify='left', values=[ 'Macho', 'Hembra'], state='readonly')
    self.cmb_sexo.place(x=30, y=320, width=200, height=60)
    self.cmb_sexo.current(0)
    self.Larete = Label(tab6, text="Arete", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    self.Larete.place(x=30, y=390)
    self.Earete=Entry(tab6, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    self.Earete.place(x=30, y=440, width=200, height=60)

####tab5###########################################################################################
    self.grid3 = ttk.Treeview(tab5, columns=('', ''), selectmode='browse')
    ####scrollbar##### 
    ##
    vsb = ttk.Scrollbar(tab5, orient="vertical", command=self.grid3.yview) 
    vsb.pack(side='right', fill='y')
    #self.grid3 = Lis(tab5, columns=('TAG', ''))
    self.grid3.place(x=-10, y=0, width=720, height=900)
    self.grid3.heading('#0', text='ID')
    self.grid3.heading('#1', text='TAG')


  # def ordenarColumna(self, id):
  #   if id == '#1':
  #       self.grid3.sort_column('#1', tk.ASCENDING)
  #       self.grid3.sort_column('#1', tk.DESCENDING)

  # def sortby(self, grid3, col, descending):
  #   # grab values to sort
  #   data = [(self.grid3.set(child, col), child) for child in self.grid3.get_children('')]
  #   # reorder data
  #   data.sort(reverse=descending)
  #   for indx, item in enumerate(data):
  #       self.grid3.move(item[1], '', indx)
  #   # switch the heading so that it will sort in the opposite direction
  #   self.grid3.heading(col, command=lambda col=col: self.grid3(grid3, col, int(not descending)))


#thisForm.list1.ListIndex = thisForm.list1.ListCount 

if __name__ == '__main__':
    ws = Tk()
    application = Principal(ws)
    ws.mainloop()