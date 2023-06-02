from tkinter import ttk
import binascii
import csv
import os
import threading
from tkinter import *
import tkinter as tk
from tkinter import messagebox
import time
import socket
import sqlite3
from tkinter import font

import openpyxl
from PIL import ImageTk, Image  

#from PIL import Image, ImageTk
#import customtkinter

class Principal():
  # customtkinter.set_appearance_mode("System")
  # customtkinter.set_default_color_theme("blue")
  db_name = 'trampa2.db'
###funcion para iniciar el socket###########################################################################################################################  
  def iniciarSocket(self):
   server_address = ('192.168.0.231', 7086)
   try:
     self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     self.sock.connect(server_address)
     self.EestadoAntena2.insert(0, 'Conectada')
   except:
     self.EestadoAntena2.insert(0, 'Desconectada')
  def __init__(self,ws):
    ##icono de la ventana 
        

   
        self.widgets()
        self.ListaCombo()
        self.iniciarSocket()
        self.IniciarLectura()
        self.VerDatos()
        
        s = ttk.Style() 
        s.theme_use('alt')
        self.respaldo=0

        s.map('TCombobox', foreground=[('readonly', 'white')])
        s.map('TCombobox', background=[('readonly', 'black')])
        s.map('TCombobox', arrowcolor=[('readonly', 'white')])
        s.map('TCombobox', selectborderwidth=[('readonly', '0')])
        s.map('TCombobox', bordercolor=[('readonly', 'black')])
        s.map('TCombobox', borderwidth=[('readonly', '0')])
        s.map('TCombobox', relief=[('readonly', 'flat')])
        s.map('TCombobox', arrowcolor=[('readonly', 'white')])
        s.map('TCombobox', selectbackground=[('readonly', 'black')])
        s.map('TCombobox', selectforeground=[('readonly', 'white')])
        s.map('TCombobox', fieldbackground=[('readonly', 'black')])
        s.map('TCombobox', selectborderwidth=[('readonly', '0')])
        #s.configure('W.TCombobox',arrowsize = 60)


        bigger_font = font.Font(family="Helvetica", size=40)
        ws.option_add('*TCombobox*Listbox.font', bigger_font)

        s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, fieldbackground= "black", background='white', foreground='black', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TNotebook', background='#003C78', borderwidth=0, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TFrame', background='#003C78', borderwidth=0 , focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TLabel', background='#003C78')

        #s.configure('TButton', background='#000000', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='white', font=('Arial','20'))       
        # s.configure("Treeview.Heading", font=(None, 28))
        # s.configure("Treeview.Heading", background="white")
        # s.configure("Treeview", font=(None, 30),  fieldbackground="black", background="black", foreground="white", selectbackground="black", selectforeground="white", arrowcolor="black", bordercolor="black", lightcolor="black", darkcolor="black", troughcolor="black", focuscolor="black", selectcolor="black")
        # s.configure("Treeview", rowheight=40,)   
        s.configure("Treeview.Heading", font=('Arial', 28))
        s.configure("Treeview.Heading", background="white")
        s.configure("Treeview", font=(None, 30),  fieldbackground="#d3eaf2", background="#d3eaf2", foreground="black", selectbackground="black", selectforeground="white", arrowcolor="black", bordercolor="black", lightcolor="black", darkcolor="black", troughcolor="black", focuscolor="black", selectcolor="black")
        s.configure("Treeview", rowheight=60, )   
        s.configure("TSeparator", background="black", foreground="black", fieldbackground="black", selectbackground="black", selectforeground="black", arrowcolor="black", bordercolor="black", lightcolor="black", darkcolor="black", troughcolor="black", focuscolor="black", selectcolor="black")

####funccion para iniciar la lectura#####################################################################################     
  def IniciarLectura(self):  
      try:    
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        self.sock.send(Read)
        hilo1=threading.Thread(target=self.GetData).start()
      except:

        return None
#######Funcion para crear la tabla en caso de no existir #############################################################################################################
  def create_database(self):
    try:
        self.conn = sqlite3.connect('trampa2.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS datos2 (id INTEGER, arete INTEGER, estatura INTEGER, peso INTEGER, fecha TEXT, hora TEXT)')
        self.conn.commit()
        messagebox.showinfo("Base de datos", "Base de datos creada con exito")
    except:
        messagebox.showerror("ERROR", "Error al crear la base de datos")
        self.conn.close()
        return None
#### Funcion para hacer la conexio###########################################################################################################################
  def conexion(self, query, parameters):
     try:
      with sqlite3.connect(self.db_name) as conn:
        cursor = conn.cursor()
        cursor.execute(query, parameters)
        
        conn.commit()
        #print("eliminado")
        cursor.close()
        
        return True 
     except Exception as e:
        #messagebox.showinfo("Base de datos", "ERROR al conectar con la base de datos")
        print(e)
        return False

######funcion para guardar los datos en la base de datos###############################################################################################            
  def GuardarEnBD(self):
    fecha = time.strftime("%d-%m-%Y")
    hora = time.strftime("%Hhrs%Mmin")
    if self.Eestatura.get() == '' or self.Epeso.get() == '':
      self.EestadoAntena2.delete(0, END) 
      self.EestadoAntena2.insert(0, 'Ingresa todos los datos')
    else:
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Conectada')
      try:
        query = 'INSERT INTO datos VALUES (?, ?, ?, ?, ?,?)'
        parameters = (self.Cidv.get(), self.Earete_entry.get(), self.Eestatura.get(), self.Epeso.get(), fecha, hora)

        self.conexion(query, parameters)

        self.EestadoAntena2.delete(0, END)
        self.EestadoAntena2.insert(0, 'OK')
        self.VerDatos()
        
        #self.conn.close()
      except Exception as e:
        messagebox.showerror(e,e)
        return None
      #self.EidV.delete(0, END)
      self.Earete_entry.delete(0, END)
      self.Eestatura.delete(0, END)
      self.Epeso.delete(0, END)
###FUNCION PARA BORRAR LOS DATOS DE LA BASE DE DATOS#############################################################################################################
  def deleteRegister(self):
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'DELETE FROM tags'
      cursor.execute(query)
      conn.commit()
      cursor.close()
    
#####fUNCION PARA GUARDAR EN XLSX###############################################################################################################
  def dbToXLSX(self):
    try:
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "Hoja1"
      ws['A1'] = 'ID' #se le asigna un titulo a la columna A
      ws['B1'] = 'UHF' #se le asigna un titulo a la columna B
      ws['C1'] = 'PESO' #se le asigna un titulo a la columna C
      ws['D1'] = 'ESTATURA' #se le asigna un titulo a la columna D
      ws['E1'] = 'FECHA' #se le asigna un titulo a la columna E
      ws['F1'] = 'HORA' #se le asigna un titulo a la columna F

      selection = self.cmb_usb.get()
      fecha = time.strftime("%d-%m-%Y")
      hora = time.strftime("%Hhrs%Mmin")

      with sqlite3.connect(self.db_name) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM datos")
        rows = cursor.fetchall()
        for row in rows:
          ws.append(row)
        
        #wb.save(selection+'atrampa2'+fecha+'-'+hora+'.xlsx')
        wb.save(selection+'\Trampa-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
        self.EestadoAntena2.delete(0, END)
        self.EestadoAntena2.insert(0, 'Datos exportados a Excel')
        cursor.close()
        #messagebox.showinfo('Base de datos', 'Datos exportados con exito')
    except Exception as e:
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Seleccione una USB valida')
      return None

######FUNCION PARA VER LOS DATOS EN LA TABLA#############################################################################################################
  def VerDatos(self):
    #self.lista2 = []
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM datos'
      result = cursor.execute(query)
      conn.commit()
      for row in result:
        #  self.lista2.append(row[1])
         self.grid3.insert ('', 0, text=row[0], values=(row[1], row[2], row[3]))
      cursor.close()

######Funcion para guardar en CVS######################################################################################################################
  def dbToExcel(self):
    try:
      conn = sqlite3.connect(self.db_name)
      cursor = conn.cursor()
      cursor.execute('SELECT * FROM datos')
      with open('datos.csv', 'w', newline='') as f:
        thewriter = csv.writer(f)
        thewriter.writerow([i[0] for i in cursor.description])
        thewriter.writerows(cursor)

      messagebox.showinfo('AGROPLUS', 'Los datos han sido exportados a Excel')
      cursor.close()
    except:
      messagebox.showerror('ERROR', 'Error al exportar los datos a Excel')
      return None

#####Funcion para obtener los datos de la antena###########################################################################################
  def GetData(self):
        self.lista = []        
        leidas = 5
        i=1
        while True:
         data = self.sock.recv(4096)  
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             if data not in self.lista :
              self.lista.append(data)
              leidas = (len(self.lista))

              if data not in self.lista:          
               self.grid3.insert('',0,text=""+str(i),values=data)                       
               i=i+1
              for column in self.grid3['columns']:                
                self.grid3.heading(column, text=column)               
            for leidas in self.lista:
                self.Earete_entry.delete(0, END)
                self.Earete_entry.insert(tk.END, data)

####funccion para detener la lectura##############################################################################################################################
  def DetenerLectura(self):   
    try:      
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
    except:
        messagebox.showerror('ERROR', 'Error al detener la lectura')
        return None

####funcion para salir##################################################################################################################################
  def salir(self): 
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
        time.sleep(1)
        self.sock.close()
        Principal1.main()

####funcion par seleccionar usb #######################################################################################################################
  def select_usb(self):
        usb = os.system('fsutil fsinfo drives')
        usb = usb.split('\n')
        listaUSB = [ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:']
        print ("cesar ",usb)
        return listaUSB

#####funcion para limpiar campos (zero) #######################################################################################################################
  def DeleteAll(self):
        self.btn.delete('1.0', END)
        self.lista = []
        self.Earete_entry.delete(0, END)
        self.Epeso.delete(0, END)

###funcion para eliminar tabla ################################################################################################################################
  def DropTable(self):
        self.conexion("DROP TABLE tags")
        messagebox.showinfo("AGROPLUS", "La tabla ha sido eliminada")
        self.VerDatos()   

####FUNCION PARA LLENAR EL CAMPO SELECCIONADO###################################################################################################################################
  def btnClick(self, num):    
         if ws.focus_get() == self.Cidv:          
        # if self.Earete_entry: 
           self.Cidv.insert(tk.END, num)
         if ws.focus_get() == self.Epeso:
           self.Epeso.insert(tk.END, num) 
         if ws.focus_get() == self.Eestatura :
           self.Eestatura.insert(tk.END, num)

#####btn limpiar ##############################################################################################################################################
  def btnClear(self):
    if ws.focus_get() == self.Cidv: 
        self.Cidv.delete(0, END)
        #self.EidV.insert('s')
    if ws.focus_get() == self.Epeso:
        self.Epeso.delete(0, END)
    if ws.focus_get() == self.Eestatura:
        self.Eestatura.delete(0, END) 
    
#####funcion para crear widgets##################################################################################################################################
  def widgets(self):
#####FRAME DE LA IZQUIERDA##################################################################################################################################
    


    left_frame = Frame(ws, bd=3, relief=SOLID, padx=210, pady=341, background='#003C78', highlightbackground="white", highlightthickness=0)
    left_frame.place(x=10, y=10)




    LidV=Label(left_frame, text="Implante", font=('ARIAL', 34), background='#003C78', foreground='#ffffff')
    LidV.place(x=-160, y=-280)

    self.Cidv = ttk.Combobox(left_frame, font=('ARIAL', 24), background='#ccd7e0', state='readonly', width=2, justify='center')
    self.Cidv.place(x=-160, y=-200, width=200, height=60)


    Larete_uhf=Label(left_frame, text="Arete UHF", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    Larete_uhf.place(x=80, y=-280)
    self.Earete_entry = Entry(left_frame, font=('ARIAL', 34), background='#ccd7e0', width=10)
    self.Earete_entry.place(x=80, y=-200, width=280, height=60)
 
    self.Lestatura = Label(left_frame, text="Estatura", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    self.Lestatura.place(x=-160, y=-20)
    self.Eestatura = Entry(left_frame, font=('ARIAL', 34), background='#070807', foreground='#00ff80')
    self.Eestatura.place(x=-160, y=50, width=200, height=60)

    self.Lpeso = Label(left_frame, text="Peso", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    self.Lpeso.place(x=80, y=-20)
    self.Epeso = Entry(left_frame, font=('ARIAL', 34), background='#070807', foreground='#00ff80')
    self.Epeso.place(x=80, y=50, width=280, height=60)
    self.Epeso.config( disabledbackground='#b8b8b8', disabledforeground='#000000', width=1)

    Euid = Entry(left_frame, font=('Arial', 24), background='#b8b8b8', justify='center' , width=1)
    Euid.place(x=-160, y=150, width=80, height=60)
   
    Euid.insert(tk.END, "7")
    EestadoAntena = Entry(left_frame, font=('ARIAL', 28), background='#ffffff', justify='center')
    EestadoAntena.place(x=-70, y=150, width=440, height=60)
    EestadoAntena.insert(0, 'Estado de la antena')
    separador = ttk.Separator(left_frame, orient=HORIZONTAL)
    separador.place(x=-160, y=225, width=530, height=3)
    self.EestadoAntena2 = Entry(left_frame, font=('ARIAL', 30), background='#070807', justify='center', foreground='#1fb1d1')
    self.EestadoAntena2.place(x=-160, y=240, width=530, height=80)
    #self.EestadoAntena2.insert(0, 'Conectada')
    login_btn = Label(left_frame, width=15, text='', font=('Times', 0), command=None, background='#003C78')
    login_btn.grid(row=10, column=1, pady=10, padx=200)


########FRAME DE LA DERECHA##############################################################################################
    right_frame = Frame(ws, bd=3,  background='#003C78', relief=SOLID, padx=320, pady=340)
    right_frame.place(x=650, y=10)
    re_pw = Label(right_frame, font=('', 14))
    re_pw.grid(row=6, column=1, pady=10, padx=20)

######pestañas############################################################################################################
    tabControl = ttk.Notebook(right_frame)
    tab1 = ttk.Frame(tabControl, width=100, height=20)
    tab5 = ttk.Frame(tabControl, width=100, height=400)
    #tab6 = ttk.Frame(tabControl, width=100, height=20)
    tabControl.add(tab1, text ='Captura')
    tabControl.add(tab5, text ='Leidas')
    #tabControl.add(tab6, text ='Registar')
    tabControl.place(x=-320, y=-340, width=687, height=590)

    #tab5.place(x=-320, y=-340, width=687, height=590)
    Bxls = Button(right_frame, text='XLS', font=('ARIAL', 24), background='#DCE1D8', command=self.dbToXLSX, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bxls.place(x=-290, y=280, width=200, height=80)
    Bsalir = Button(right_frame, text='X', font=('ARIAL', 24), background='red', command=self.salir, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bsalir.place(x=205, y=-337, width=80, height=80)

    self.btnguardar = Button(right_frame, text="Guardar", font=('ARIAL', 18), background='#DCE1D8', command=self.GuardarEnBD, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    self.btnguardar.place(x=100, y=280, width=200, height=80)
    self.cmb_usb=ttk.Combobox(right_frame, width=10, font=('ARIAL', 28), justify='left', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:'], state='readonly')
    self.cmb_usb.place(x=-40, y=280, width=100, height=60)
    self.cmb_usb.current(1)
    # separador2 = ttk.Separator(tab1, orient=HORIZONTAL, style='TSeparator')
    # separador2.place(x=0, y=0, width=800, height=4)
####tab1########################################################################################################################################################################
    self.btn1 = tk.Button(tab1, text="1",  width=6, height=2, command=lambda:self.btnClick(1),  border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8',)
    self.btn1.place(x=105, y=35)
    self.btn2 = tk.Button(tab1, text="2", width=6, height=2, command=lambda:self.btnClick(2), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn2.place(x=250, y=35)
    self.btn3 = tk.Button(tab1, text="3", width=6, height=2, command=lambda:self.btnClick(3), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn3.place(x=395, y=35)
    self.btn4 = tk.Button(tab1, text="4", width=6, height=2, command=lambda:self.btnClick(4), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn4.place(x=105, y=155)
    self.btn5 = tk.Button(tab1, text="5", width=6, height=2, command=lambda:self.btnClick(5), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn5.place(x=250, y=155)
    self.btn6 = tk.Button(tab1, text="6", width=6, height=2, command=lambda:self.btnClick(6), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn6.place(x=395, y=155)
    self.btn7 = tk.Button(tab1, text="7", width=6, height=2, command=lambda:self.btnClick(7), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn7.place(x=105, y=275)
    self.btn8 = tk.Button(tab1, text="8", width=6, height=2, command=lambda:self.btnClick(8), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn8.place(x=250, y=275)
    self.btn9 = tk.Button(tab1, text="9", width=6, height=2, command=lambda:self.btnClick(9), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn9.place(x=395, y=275)
    self.btn0 = tk.Button(tab1, text="0", width=6, height=2, command=lambda:self.btnClick(0), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn0.place(x=250, y=395)
    self.btnborrar = tk.Button(tab1, text="C", width=6, height=2, command=self.btnClear, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btnborrar.place(x=105, y=395)
    self.btnborrar = tk.Button(tab1, text="ID", width=6, height=2, command=lambda:self.btnClick(""), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btnborrar.place(x=395, y=395)
    


####tab5####################################################################################################################
    self.grid3 = ttk.Treeview(tab5, columns=('', '', ''), selectmode='browse')       
    ####scrollbar##### 
    vsb = ttk.Scrollbar(tab5, orient="vertical", command=self.grid3.yview) 
    #vsb.place(x=-700, y=-100, height=500)
    vsb.pack(side='right', fill='y')
    #self.grid3 = Lis(tab5, columns=('TAG', ''))
    self.grid3.column('#0', width=178, stretch=NO)
    self.grid3.column('#1', width=200, stretch=NO)
    self.grid3.column('#2', width=150, stretch=NO)
    self.grid3.column('#3', width=150, stretch=NO)

    self.grid3.pack(side='left', fill='both', expand=True)
    self.grid3.heading('#0', text='Inplante', anchor='center')
    self.grid3.heading('#1', text='UHF')
    self.grid3.heading('#2', text='Peso')
    self.grid3.heading('#3', text='Estatura')
#########tab6############################################################################################################################################
    # self.Limplante = Label(tab6, text="Implantes", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    # self.Limplante.place(x=30, y=30)
    # # self.Eimplante=Entry(tab6, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    # # self.Eimplante.place(x=30, y=80, width=200, height=60)
    # self.Btnguardar2 = Button(tab6, text="Guardar", font=('ARIAL', 18), background='#DCE1D8', command=self.guardarImplante, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    # self.Btnguardar2.place(x=560, y=35, width=100, height=120)

    # self.Limplante = Label(tab6, text="Implantes", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    # self.Limplante.place(x=30, y=30)
    # self.Eimplante=Entry(tab6, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    # self.Eimplante.place(x=30, y=80, width=200, height=60)
    # self.Btnguardar2 = Button(tab6, text="Guardar", font=('ARIAL', 18), background='#DCE1D8', command=self.guardarImplante, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    # self.Btnguardar2.place(x=560, y=35, width=100, height=120)

######ventana de registro ############################################################################################################
  def VentanaRegistro(self):
   
    s = ttk.Style()
    s.theme_use('alt')

    
    s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, fieldbackground= "black", background='#d3eaf2', 
    foreground='black', borderwidth=3, relief='groove', focuscolor="", selectcolor='black', 
    selectbackground='black', selectforeground='black', bordercolor='black', activebackground='black',  
    activeforeground='black', highlightcolor='black', highlightbackground='black', highlightthickness=3)

    s.configure('TNotebook', background='#a8d5e5', borderwidth=0, focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black', )
    
    s.configure('TFrame', background='#a8d5e5', borderwidth=2 , focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black',  )
    
    s.configure('TLabel', background='#003C78')
    s.configure("Treeview.Heading", font=('Arial', 28))
    s.configure("Treeview.Heading", background="#d3eaf2", foreground="black", fieldbackground="#a8d5e5")
    s.configure("Treeview", font=(None, 30),  fieldbackground="white", background="white",
    foreground="black", selectbackground="black", selectforeground="white", arrowcolor="black", 
    bordercolor="black", lightcolor="black", darkcolor="black", troughcolor="black", focuscolor="black", 
    selectcolor="black", borderwidth=3, relief='sunken')
    s.configure("Treeview", rowheight=50,  )   

    self.VentanaRegistro = Toplevel(ws)
  
    #self.VentanaRegistro.title('Registrar')
    self.VentanaRegistro.geometry('1400x900')
    # self.VentanaRegistro.resizable(0, 0)
    # self.VentanaRegistro.transient(self .ws)

    # self.VentanaRegistro.grab_set()
    # self.VentanaRegistro.focus_set( )
    self.VentanaRegistro.configure(background='white',) #bee0ec azul

    tabRegistro = ttk.Notebook(self.VentanaRegistro)


    tab11 = ttk.Frame(tabRegistro, width=1300, height=700)
    #tab11.configure(width=1000, height=200)
    tab12 = ttk.Frame(tabRegistro, width=100, height=400)
    tab13 = ttk.Frame(tabRegistro, width=100, height=20)
    tab14 = ttk.Frame(tabRegistro, width=100, height=20)
    tab15 = ttk.Frame(tabRegistro, width=100, height=20)
    tab16 = ttk.Frame(tabRegistro, width=100, height=20)
    tab17 = ttk.Frame(tabRegistro, width=100, height=20)
    tabRegistro.add(tab11, text ='Implantes')
    tabRegistro.add(tab12, text ='Vendedores')
    tabRegistro.add(tab13, text ='Registros')
    tabRegistro.add(tab14, text ='Reportes')
    tabRegistro.add(tab15, text ='Configuración')
    tabRegistro.add(tab16, text ='Ayuda')
    tabRegistro.add(tab17, text ='        Salir      ')
    tabRegistro.place(x=0, y=0, width=1500, height=740, )
    
    #tabRegistro.configure(width=1887, height=590, )

    self.gridRegistro = ttk.Treeview(tab11, columns=('',), selectmode='browse', height=50,  )
   
    ####scrollbar##### 
    #vsb = ttk.Scrollbar(tab11, orient="vertical", command=self.gridRegistro.yview) 
    #vsb.place(x=-700, y=-100, height=500)
    #vsb.pack(side='right', fill='y')
    #self.grid3 = Lis(tab5, columns=('TAG', ''))
    self.gridRegistro.column('#0', width=178, stretch=NO, anchor='center') 
    self.gridRegistro.column('#1', width=420, stretch=NO, anchor="center" )
    
    #self.gridRegistro.pack(side='left', fill='both',  padx=20, pady=20,ipadx=20, ipady=20)
    self.gridRegistro.place(x=20, y=60, width=600, height=550, )
    self.gridRegistro.heading('#0', text='N°.', anchor='center')
    self.gridRegistro.heading('#1', text='IMPLANTE')
    self.gridRegistro.bind('<ButtonRelease-1>', self.selectItem)

    # self.Limplante = Label(tab11, text="Implantes", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    # self.Limplante.place(x=700, y=30)
    self.EestadoRegistro = Entry(tab11, font=('ARIAL', 28), background='#d3eaf2', justify='center', foreground='black',  border='3', borderwidth=3, relief='ridge')
    self.EestadoRegistro.place(x=700, y=65, width=500, height=60)
    self.EestadoRegistro.insert(0, '')

    self.Eimplante=Entry(tab11, font=('ARIAL', 28), background='white', justify='center', foreground='black', relief='ridge', borderwidth=2  ) 
    self.Eimplante.place(x=800, y=300, width=300, height=60)

    # separador1 = ttk.Separator(self.VentanaRegistro, orient=HORIZONTAL, style='TSeparator')
    # separador1.place(x=700, y=580, width=500, height=1)
    # separador2 = ttk.Separator(self.VentanaRegistro, orient=HORIZONTAL, style='TSeparator')
    # separador2.place(x=700, y=100, width=0, height=1)
    # separador3 = ttk.Separator(self.VentanaRegistro, orient=VERTICAL, style='TSeparator')
    # separador3.place(x=700, y=100, width=1, height=480)
    # separador4 = ttk.Separator(self.VentanaRegistro, orient=VERTICAL, style='TSeparator')
    # separador4.place(x=1200, y=100, width=1, height=480)

    #self.EestadoRegistro.pack(side='right', fill='both',  padx=10, pady=10)
    self.BaceptarRegistro = Button(tab11, text="Aceptar", font=('ARIAL', 18), background='#DCE1D8', command=self.AceptarRegistro, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    self.BaceptarRegistro.place(x=700, y=530, width=200, height=80)
    self.BtnSalir2 = Button(tab11, text="Salir", font=('ARIAL', 18), background='RED', command=self.VentanaRegistro.destroy, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    self.BtnSalir2.place(x=1000, y=530, width=200, height=80)

#icono eliminar

    # path22 = "trash.jpg"
    # img22 = ImageTk.PhotoImage(Image.open(path22))
    # panel22 = Label(self.ws, image=img22, bg='black')
    # panel22.photo = img22
    # panel22.place(x=0, y=0)
    # label=Label(self.VentanaRegistro, text="Eliminar", font=('ARIAL', 18), background='black', foreground='white', justify='center', image=img22, compound=LEFT)
    # label.place(x=0, y=0)

    self.radio = IntVar()
    # self.Rbutton1 = Radiobutton(tab11, text="Añadir", variable=self.radio, command=self.Selection, value=1, background='#a8d5e5', foreground='black', font=('ARIAL', 24))
    # #self.Rbutton1 = customtkinter.CTkRadioButton(self.VentanaRegistro, text="Añadir", command=self.Selection, value=1, font=('ARIAL', 18))
    # self.Rbutton1.place(x=1040, y=200)
    # self.Rbutton2 = Radiobutton(tab11, text="Modificar",  variable=self.radio,command=self.Selection, value=2, background='#a8d5e5', foreground='black', font=('ARIAL', 24))
    # self.Rbutton2.place(x=1040, y=250)
    # self.Rbutton3 = Radiobutton(tab11, text="Eliminar", variable=self.radio, command=self.Selection, value=3, background='#a8d5e5', foreground='black', font=('ARIAL', 24))
    # self.Rbutton3.place(x=1040, y=300)
    self.Rbutton1 = Radiobutton(tab11, text="Añadir", variable=self.radio, command=self.Selection, value=1, background='#a8d5e5', foreground='black', font=('ARIAL', 24))
    #self.Rbutton1 = customtkinter.CTkRadioButton(self.VentanaRegistro, text="Añadir", command=self.Selection, value=1, font=('ARIAL', 18))
    
    self.Rbutton1.place(x=700, y=200)
    self.Rbutton2 = Radiobutton(tab11, text="Modificar",  variable=self.radio,command=self.Selection, value=2, background='#a8d5e5', foreground='black', font=('ARIAL', 24))
    self.Rbutton2.place(x=860, y=200)
    self.Rbutton3 = Radiobutton(tab11, text="Eliminar", variable=self.radio, command=self.Selection, value=3, background='#a8d5e5', foreground='black', font=('ARIAL', 24))
    self.Rbutton3.place(x=1040, y=200)
    ##icono eliminar
    # label=Label(self.VentanaRegistro, text="", font=('ARIAL', 18), background='white', foreground='white', justify='center', image=img22, compound=LEFT, border=0)
    # label.place(x=750, y=395)

    self.VerImplantes()

#####evento de click en la tabla de implantes###############################################################################
  def selectItem(self, event):
    curItem = self.gridRegistro.item(self.gridRegistro.focus()) #obtiene el item seleccionado
    col = self.gridRegistro.identify_column(event.x)
   
    if col == '#0':
        cell_value = curItem['values'][0]
    elif col == '#1':
        cell_value = curItem['values'][0]
    elif col == '#2':
        cell_value = curItem['values'][0]

    print ('cell_value = ', cell_value)
    self.Eimplante.delete(0, END)
    self.Eimplante.insert(0, cell_value)

###########btn de accion en registro ############################################################################################################
  def AceptarRegistro(self):
    if self.radio.get() == 1:    
      self.guardarImplante()
      self.Eimplante.delete(0, END)
    elif self.radio.get() == 2:
      self.UpdateImplante()
      self.Eimplante.delete(0, END)
    elif self.radio.get() == 3:
      self.BorrarImplante()
      self.Eimplante.delete(0, END)
    else:
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Seleccione una opción')

####funcion para los rb de registro###############################################################################################################
  def Selection(self):
    print (self.radio.get())
    if self.radio.get() == 1:
      #relese focus of the treeview
      self.Rbutton1.config(foreground='green')
      self.Rbutton2.config(foreground='black')
      self.Rbutton3.config(foreground='black')
      self.gridRegistro.selection_remove(self.gridRegistro.focus())
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Ingresa un implante')
      #self.Eimplante.delete(0, END)
      #self.Eimplante.focus()
    elif self.radio.get() == 2:
      self.Rbutton1.config(foreground='black')
      self.Rbutton2.config(foreground='green')
      self.Rbutton3.config(foreground='black')
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Selecciona un implante')
      
    elif self.radio.get() == 3:
      self.Rbutton1.config(foreground='black')
      self.Rbutton2.config(foreground='black')
      self.Rbutton3.config(foreground='green')
      self.Eimplante.delete(0, END)
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Selecciona un implante') 
    else:
      self.Rbutton1.config(foreground='black')
      self.Rbutton2.config(foreground='black')
      self.Rbutton3.config(foreground='black')       

########### funcion para guardar implante en la base de datos############################################################################################################
  def guardarImplante(self):
    if self.Eimplante.get() == '':
        #self.EestadoRegistro.delete(0, END)
        self.EestadoRegistro.insert(0, 'Ingresa un implante')
      
    else:
      try:

        query = 'INSERT INTO implantes VALUES (?, ?)'
        parameters = ((self.idimp+1), self.Eimplante.get())
        #self.conexion(query, parameters)
        print(query, parameters )
        if self.conexion(query, parameters):
          self.VerImplantes()
          self.EestadoRegistro.delete(0, END)
          self.EestadoRegistro.insert(0, 'Implante guardado')
          
        else:
          self.VerImplantes()
          self.EestadoRegistro.delete(0, END)
          self.EestadoRegistro.insert(0, 'El implante ya existe')
   
      except:    
        self.EestadoRegistro.delete(0, END)
        self.EestadoRegistro.insert(0, 'ERROR AL GUARDAR')

###### Borra el implante seleccionado de la base de datos ############################################################################################################
  def BorrarImplante(self):
   self.conte=0
   if self.gridRegistro.selection():
    try:
      query = 'DELETE FROM implantes WHERE implante=?'
      parameters = (self.gridRegistro.item(self.gridRegistro.selection())['values'][0])
      print  (self.gridRegistro.item(self.gridRegistro.selection())['values'][0])
      print(query, (parameters,))
      self.conexion(query, (parameters,))
      self.VerImplantes()
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Implante eliminado')
      self.radio.set(0)
      self.conte += 1
      self.Rbutton3.config(foreground='black')  
    except:
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Error al eliminar')
   else:
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Selecciona un implante')
  
  
#cargar los implantes en la tabla ############################################################################################################
  def VerImplantes(self):
    self.EestadoRegistro.insert(0, 'Selecciona una opción')
    self.idimp=0
    records = self.gridRegistro.get_children()     #cleaning table
    for element in records:
            self.gridRegistro.delete(element)
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM implantes ORDER BY id DESC'
      result = cursor.execute(query)
      conn.commit()
      for row in result:

         self.gridRegistro.insert ('', 0, text=row[0], values=row[1])
         self.idimp=self.idimp+1

      cursor.close()

########### funcion para actualizar el implante seleccionado en la base de datos############################################################################################################
  def UpdateImplante(self):
   if self.Eimplante.get() != '' and self.gridRegistro.selection():
     try:
      query = 'UPDATE implantes SET implante=? WHERE id=? AND implante=?'
      parameters = (self.Eimplante.get(), self.gridRegistro.item(self.gridRegistro.selection())['text'], self.gridRegistro.item(self.gridRegistro.selection())['values'][0])
      print (parameters)
      self.conexion(query, parameters)
      self.VerImplantes()
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Registro actualizado')
     except Exception as e:
        self.EestadoRegistro.delete(0, END)
        self.EestadoRegistro.insert(0, 'Selecciona un implante')
   else:
      self.EestadoRegistro.delete(0, END)
      self.EestadoRegistro.insert(0, 'Selecciona un implante')    

########### funcion para cargar los datos del implante seleccionado en el combo box############################################################################################################
  def ListaCombo(self):
    self.lista3 = []
    
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM implantes'
      result = cursor.execute(query)
      conn.commit()
      for row in result:
         self.lista3.append(row[1])
         self.Cidv['values'] = self.lista3
      cursor.close()


  def main():
    if __name__ == '__main__':
      
      Principal(ws)

################################################################################################################################################################
################################################################################################################################################################
################################################################################################################################################################
################################################################################################################################################################

#####MENU PRINCIPAL############################################################################################################################################################################
class Principal1(Principal):
       
  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus Menu')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='#003C78')
        
        self.widgets()

#####funcion para crear widgets##################################################################################################################################
  def acerca_de(self):
        self.ventana = Toplevel(self.ws)
        self.ventana.title('Acerca de')
        self.ventana.geometry('400x300')

        self.ventana.iconbitmap('ico.ico')

        self.ventana.resizable(0, 0)
        self.ventana.transient(self.ws)
        # self.ventana.grab_set()
        # self.ventana.focus_set( )
        self.ventana.configure(background='#003C78',)

        #posiciona la ventana en el centro de la pantalla
        x = self.ws.winfo_x()
        y = self.ws.winfo_y()
        w = self.ws.winfo_width()
        h = self.ws.winfo_height()
        self.ventana.geometry("+%d+%d" % (x + w/2 - 200, y + h/2 - 150))


        # self.ventana.attributes("-topmost", True)
        # self.ventana.attributes("-toolwindow", True)


        self.texto = Label(self.ventana, text='Agroplus', font=('Arial', 20), bg='#003C78', fg='white')
        self.texto.place(x=150, y=20)

        self.texto = Label(self.ventana, text='Version 1.0', font=('Arial', 12), bg='#003C78', fg='white')
        self.texto.place(x=120, y=80)

        self.texto = Label(self.ventana, text='Desarrollado por: ', font=('Arial', 12), bg='#003C78', fg='white')
        self.texto.place(x=120, y=120)

        self.texto = Label(self.ventana, text='Ing. César Armando Del Río', font=('Arial', 12), bg='#003C78', fg='white')
        self.texto.place(x=120, y=140)

        self.texto = Label(self.ventana, text='Ing. Raymundo Rodriguez', font=('Arial', 12), bg='#003C78', fg='white')
        self.texto.place(x=120, y=170)

        self.btnExit=Button(self.ventana, text='Salir', command=self.ventana.destroy, bg='white', fg='black', font=('Arial', 12))
        self.btnExit.place(x=150, y=220)

  def widgets(self):

        # fondo.place(x=90, y=90)
        self.frame1 = Frame(self.ws, bg='black')
        self.frame1.place(x=0, y=0, width=1366, height=768)

        path = "BOVYFOX.png"
        img = ImageTk.PhotoImage(Image.open(path))
        panel = Label(self.ws, image=img, bg='black')
        panel.photo = img
        panel.place(x=370, y=0)

        btn1=self.boton = Button(self.ws, text='Inventario', command=self.ws.destroy)
        btn1.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        btn1.place(x=140, y=150, width=360, height=180)

        btn2=self.boton = Button(self.ws, text='Prensa', command=self.cerrarVentana)    
        btn2.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        btn2.place(x=140, y=340 , width=360, height=180)

        btn3=self.boton = Button(self.ws, text='Establo', command=self.ws.destroy)
        btn3.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        btn3.place(x=140, y=530, width=360, height=180)  

        btn4=self.boton = Button(self.ws, text='Buscar', command=self.ws.destroy)
        btn4.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        btn4.place(x=508, y=150, width=355, height=180)

        btn5=self.boton = Button(self.ws, text='', command=self.ws.destroy)
        btn5.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        btn5.place(x=508, y=340, width=355, height=180)

        btn6=self.boton = Button(self.ws, text='Licencia', command=self.ws.destroy)
        btn6.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        btn6.place(x=508, y=530, width=355, height=180)

        btn7=self.boton = Button(self.ws, text='Acerca de', command=self.acerca_de)
        btn7.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        btn7.place(x=872, y=150, width=355, height=180)

        btn8=self.boton = Button(self.ws, text='Registrar', command=self.VentanaRegistro)
        btn8.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        btn8.place(x=872, y=340, width=355, height=180)

        btn9=self.boton = Button(self.ws, text='Salir', command=self.ws.destroy)
        btn9.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        btn9.place(x=872, y=530, width=355, height=180)
       
        # path2 = "logo.png"
        # imagen2 = ImageTk.PhotoImage(Image.open(path2))
        # panel2 = Label(self.ws, image=imagen2, bg='black')
        # panel2.photo = img
        # panel2.place(x=350, y=100)

        # btn1=self.boton = Button(self.ws, text='Inventario', command=self.ws.destroy)
        # btn1.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        # btn1.place(x=140, y=150)

        # btn2=self.boton = Button(self.ws, text='Prensa', command=self.cerrarVentana)    
        # btn2.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        # btn2.place(x=140, y=340)

        # btn3=self.boton = Button(self.ws, text='Establo', command=self.ws.destroy)
        # btn3.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        # btn3.place(x=140, y=530  )

        # btn4=self.boton = Button(self.ws, text='Buscar', command=self.ws.destroy)
        # btn4.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        # btn4.place(x=508, y=150)

        # btn5=self.boton = Button(self.ws, text='', command=self.ws.destroy)
        # btn5.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        # btn5.place(x=508, y=340)

        # btn6=self.boton = Button(self.ws, text='', command=self.ws.destroy)
        # btn6.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        # btn6.place(x=508, y=530)

        # btn7=self.boton = Button(self.ws, text='Acerca de', command=self.ws.destroy)
        # btn7.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        # btn7.place(x=872, y=150)

        # btn8=self.boton = Button(self.ws, text='Registrar', command=self.VentanaRegistro)
        # btn8.config(width=10, height=2, font=('Arial', 48), bg='#2596be', fg='black')
        # btn8.place(x=872, y=340)

        # btn9=self.boton = Button(self.ws, text='Salir', command=self.ws.destroy)
        # btn9.config(width=10, height=2, font=('Arial', 48), bg='#a8d5e5', fg='black')
        # btn9.place(x=872, y=530)

        #imagen= PhotoImage(file="1.jpg")
        #fondo=Label(self.ws, image=imagen, text = "Imagen S.O de fondo",  ).grid(column=2,row=2)
        #fondo.photo=imagen
        #fondo.place(x=0, y=0)

  def cerrarVentana(self):
        Principal.main()
  def VentanaRegistro(self):
        Principal.VentanaRegistro(self)
  def main():
    if __name__ == '__main__':
      
      Principal1(ws)

class Menu(Principal1):
  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='black')
        self.widgets()
        s = ttk.Style()

  def widgets(self):

        path = "BOVYFOX.png"
        img = ImageTk.PhotoImage(Image.open(path))
        panel = Label(self.ws, image=img, bg='black')
        panel.photo = img
        panel.place(x=370, y=300)
       
        BtnEntrar2 = Button(self.ws, text="Entrar", font=('ARIAL', 18), background='white', command=self.pantallaMenu, border=6, )
        BtnEntrar2.place(x=1000, y=670, width=200, height=80)
        
  def pantallaMenu(self):
     Principal1.main()
     #return super().cerrarVentana()
#####FRAME DE LA IZQUIERDA##################################################################################################################################
        
if __name__ == '__main__':
    ws = Tk()
    application = Menu(ws)
    ws.mainloop()
