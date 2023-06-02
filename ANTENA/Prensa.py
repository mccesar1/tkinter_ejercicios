import binascii
import csv
from datetime import date
import datetime
from msilib.schema import ComboBox
import os
import threading
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
import time
import socket
import os
import sqlite3
import openpyxl
class Principal:
  db_name = 'trampa2.db'
###funcion para iniciar el socket###########################################################################################################################  
  def iniciarSocket(self):
   server_address = ('192.168.0.231', 7086)
   try:
     self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     self.sock.connect(server_address)
     self.EestadoAntena2.insert(0, 'Conectada')
   except:
     self.EestadoAntena2.insert(0, 'Desconectada')
  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='#003C78')
        # print (date.today().year)
        # #print (date.year)
        # if date.today().year > 2021:
        #     #introducir licencia

        #      #messagebox.QUESTION("ERROR", "Licencia vencida")
        #      #messagebox.showerror('ERROR', 'Licencia expirada')
        #      self.ws.destroy()

            
        # else:
        #   pass
        
        self.widgets()

        self.ListaCombo()

        self.iniciarSocket()
        #self.IniciarLectura()
        self.VerDatos()
        s = ttk.Style()
        s.theme_use('alt')
        self.respaldo=0
        s.map('TCombobox', fieldbackground=[('readonly','black')])
        s.map('TCombobox', selectbackground=[('readonly', 'black')])
        s.map('TCombobox', selectforeground=[('readonly', 'white')])
        s.map('TCombobox', foreground=[('readonly', 'white')])
        s.map('TCombobox', background=[('readonly', 'black')])
        s.map('TCombobox', arrowcolor=[('readonly', 'white')])
        s.map('TCombobox', selectborderwidth=[('readonly', '0')])
        s.map('TCombobox', bordercolor=[('readonly', 'black')])
        s.map('TCombobox', borderwidth=[('readonly', '0')])
        s.map('TCombobox', relief=[('readonly', 'flat')])
        s.configure("TCombobox", fieldbackground= "black", background= "black", foreground= "black", selectbackground= "black", selectforeground= "black", arrowcolor= "black", bordercolor= "black", lightcolor= "black", darkcolor= "black", troughcolor= "black", focuscolor= "black", selectcolor= "black")
        s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, fieldbackground= "black", background='black', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TNotebook', background='#003C78', borderwidth=0, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TFrame', background='#003C78', borderwidth=0 , focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        s.configure('TLabel', background='#003C78')
        #s.configure('TButton', background='#000000', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='white', font=('Arial','20'))       
        s.configure("Treeview.Heading", font=(None, 28))
        s.configure("Treeview.Heading", background="white")
        s.configure("Treeview", font=(None, 30),  fieldbackground="black", background="black", foreground="white", selectbackground="black", selectforeground="white", arrowcolor="black", bordercolor="black", lightcolor="black", darkcolor="black", troughcolor="black", focuscolor="black", selectcolor="black")
        s.configure("Treeview", rowheight=40,)   
          
####funccion para iniciar la lectura#####################################################################################     
  def IniciarLectura(self):  
      try:    
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        self.sock.send(Read)
        hilo1=threading.Thread(target=self.GetData).start()
      except:
        #self.EestadoAntena2.insert(0, 'Desconectada')
        return None
#######Funcion para crear la tabla en caso de no existir #############################################################################################################
  def create_database(self):
    try:
        self.conn = sqlite3.connect('trampa2.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('CREATE TABLE IF NOT EXISTS datos2 (id INTEGER, arete INTEGER, estatura INTEGER, peso INTEGER, fecha TEXT, hora TEXT)')
        self.conn.commit()
        messagebox.showinfo("Base de datos", "Base de datos creada con exito")
    except:
        messagebox.showerror("ERROR", "Error al crear la base de datos")
        self.conn.close()
        return None
#### Funcion para hacer la conexio###########################################################################################################################
  def conexion(self, query, parameters):
     try:
      with sqlite3.connect(self.db_name) as conn:
        cursor = conn.cursor()
        cursor.execute(query, parameters)
        #cursor.execute(query, (parameters,))
        conn.commit()
        
     except Exception as e:
        messagebox.showerror("ERROR", e)
        print(e)
        return None

 
        #messagebox.showinfo("Base de datos", "Base de datos cerrada con exito")
######funcion para guardar los datos en la base de datos###############################################################################################            
  def GuardarEnBD(self):
    fecha = time.strftime("%d-%m-%Y")
    hora = time.strftime("%Hhrs%Mmin")
    if self.Eestatura.get() == '' or self.Epeso.get() == '':
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Ingresa todos los datos')
    else:
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Conectada')
      try:
        query = 'INSERT INTO datos VALUES (?, ?, ?, ?, ?,?)'
        parameters = (self.Cidv.get(), self.Earete_entry.get(), self.Eestatura.get(), self.Epeso.get(), fecha, hora)
            #parameters2 = (self.grid3.item(item)['values'][1])
            #print (self.grid3.item(item)['values'][1])
        self.conexion(query, parameters)
        #self.conexion(query, (parameters,))
       
        
        self.EestadoAntena2.delete(0, END)
        self.EestadoAntena2.insert(0, 'OK')
        self.VerDatos()
        
        #self.conn.close()
      except Exception as e:
        messagebox.showerror(e,e)
        return None
      #self.EidV.delete(0, END)
      self.Earete_entry.delete(0, END)
      self.Eestatura.delete(0, END)
      self.Epeso.delete(0, END)
###FUNCION PARA BORRAR LOS DATOS DE LA BASE DE DATOS#############################################################################################################
  def deleteRegister(self):
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'DELETE FROM tags'
      cursor.execute(query)
      conn.commit()
      cursor.close()
    

#####fUNCION PARA GUARDAR EN XLSX###############################################################################################################
  def dbToXLSX(self):
    try:
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "Hoja1"
      ws['A1'] = 'ID' #se le asigna un titulo a la columna A
      ws['B1'] = 'UHF' #se le asigna un titulo a la columna B
      ws['C1'] = 'PESO' #se le asigna un titulo a la columna C
      ws['D1'] = 'ESTATURA' #se le asigna un titulo a la columna D
      ws['E1'] = 'FECHA' #se le asigna un titulo a la columna E
      ws['F1'] = 'HORA' #se le asigna un titulo a la columna F
      selection = self.cmb_usb.get()
      fecha = time.strftime("%d-%m-%Y")
      hora = time.strftime("%Hhrs%Mmin")

      with sqlite3.connect(self.db_name) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM datos")
        rows = cursor.fetchall()
        for row in rows:
          ws.append(row)
        
        #wb.save(selection+'atrampa2'+fecha+'-'+hora+'.xlsx')
        wb.save(selection+'\Trampa-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
        self.EestadoAntena2.delete(0, END)
        self.EestadoAntena2.insert(0, 'Datos exportados a Excel')
        cursor.close()
        #messagebox.showinfo('Base de datos', 'Datos exportados con exito')
    except Exception as e:
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Seleccione una USB valida')
      return None

######FUNCION PARA VER LOS DATOS EN LA TABLA#############################################################################################################
  def VerDatos(self):
    self.lista2 = []
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM datos'
      result = cursor.execute(query)
      conn.commit()
      for row in result:
        #  self.lista2.append(row[1])
         self.grid3.insert ('', 0, text=row[0], values=(row[1], row[2], row[3]))
      cursor.close()

######Funcion para guardar en CVS######################################################################################################################
  def dbToExcel(self):
    try:
      conn = sqlite3.connect(self.db_name)
      cursor = conn.cursor()
      cursor.execute('SELECT * FROM datos')
      with open('datos.csv', 'w', newline='') as f:
        thewriter = csv.writer(f)
        thewriter.writerow([i[0] for i in cursor.description])
        thewriter.writerows(cursor)

      messagebox.showinfo('AGROPLUS', 'Los datos han sido exportados a Excel')
      cursor.close()
    except:
      messagebox.showerror('ERROR', 'Error al exportar los datos a Excel')
      return None
      #self.VerDatos()
#####Funcion para obtener los datos de la antena###########################################################################################
  def GetData(self):
        self.lista = []        
        leidas = 5
        i=1
        while True:
         data = self.sock.recv(4096)  
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             if data not in self.lista :
              self.lista.append(data)
              leidas = (len(self.lista))

              if data not in self.lista:          
               self.grid3.insert('',0,text=""+str(i),values=data)                       
               i=i+1
              for column in self.grid3['columns']:                
                self.grid3.heading(column, text=column)               
            for leidas in self.lista:
                self.Earete_entry.delete(0, END)
                self.Earete_entry.insert(tk.END, data)

####funccion para detener la lectura##############################################################################################################################
  def DetenerLectura(self):   
    try:      
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
    except:
        messagebox.showerror('ERROR', 'Error al detener la lectura')
        return None
       
####funcion para salir##################################################################################################################################
  def salir(self):
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
        time.sleep(1)
        self.sock.close()

        #self.ventana.destroy()
        #sys.exit()
        #self.DetenerLectura()
        self.ws.destroy()

  def select_usb(self):
        usb = os.system('fsutil fsinfo drives')
        usb = usb.split('\n')
        listaUSB = [ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:']
        print ("cesar ",usb)
        return listaUSB

  def DeleteAll(self):
        self.btn.delete('1.0', END)
        self.lista = []
        self.Earete_entry.delete(0, END)
        self.Epeso.delete(0, END)

  def DropTable(self):
        self.conexion("DROP TABLE tags")
        messagebox.showinfo("AGROPLUS", "La tabla ha sido eliminada")
        self.VerDatos()   


  def btnClick(self, num):    
         if ws.focus_get() == self.Cidv:          
        # if self.Earete_entry: 
           self.Cidv.insert(tk.END, num)
         if ws.focus_get() == self.Epeso:
           self.Epeso.insert(tk.END, num) 
         if ws.focus_get() == self.Eestatura :
           self.Eestatura.insert(tk.END, num)

  def btnClear(self):
    if ws.focus_get() == self.Cidv: 
        self.Cidv.delete(0, END)
        #self.EidV.insert('s')
    if ws.focus_get() == self.Epeso:
        self.Epeso.delete(0, END)
    if ws.focus_get() == self.Eestatura:
        self.Eestatura.delete(0, END) 
    
#####funcion para crear widgets##################################################################################################################################
  def widgets(self):
#####FRAME DE LA IZQUIERDA##################################################################################################################################
    left_frame = Frame(ws, bd=3, relief=SOLID, padx=210, pady=340, background='#003C78', highlightbackground="white", highlightthickness=3)
    left_frame.place(x=10, y=10)

    LidV=Label(left_frame, text="Implante", font=('ARIAL', 34), background='#003C78', foreground='#ffffff')
    LidV.place(x=-160, y=-280)
   # query = 'SELECT * FROM implantes'
   # parameters = 
    #self.conexion(query, parameters)
    self.Cidv = ttk.Combobox(left_frame, font=('ARIAL', 34), background='#ccd7e0', state='readonly')
    self.Cidv.place(x=-160, y=-200, width=200, height=60)
  #self.Cidv['values'] = self.cursor.fetchall()
    # self.EidV = Entry(left_frame, font=('ARIAL', 34), background='#ccd7e0')
    # self.EidV.place(x=-160, y=-200, width=200, height=60)
    # self.Cidv = ttk.Combobox(left_frame, font=('ARIAL', 34), background='#ccd7e0', state='readonly')
    # self.Cidv.place(x=-160, y=-200, width=200, height=60)
    # self.Cidv['values'] = ()


    Larete_uhf=Label(left_frame, text="Arete UHF", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    Larete_uhf.place(x=80, y=-280)
    self.Earete_entry = Entry(left_frame, font=('ARIAL', 34), background='#ccd7e0', width=10)
    self.Earete_entry.place(x=80, y=-200, width=280, height=60)
 
    self.Lestatura = Label(left_frame, text="Estatura", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    self.Lestatura.place(x=-160, y=-20)
    self.Eestatura = Entry(left_frame, font=('ARIAL', 34), background='#070807', foreground='#00ff80')
    self.Eestatura.place(x=-160, y=50, width=200, height=60)

    self.Lpeso = Label(left_frame, text="Peso", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    self.Lpeso.place(x=80, y=-20)
    self.Epeso = Entry(left_frame, font=('ARIAL', 34), background='#070807', foreground='#00ff80')
    self.Epeso.place(x=80, y=50, width=280, height=60)
    self.Epeso.config( disabledbackground='#b8b8b8', disabledforeground='#000000', width=1)

    # self.Limplante = Label(left_frame, text="Implante", font=('ARIAL', 30), background='#003C78', foreground='#ffffff')
    # self.Limplante.place(x=-160, y=-60)
    #self.Eimplante = ComboBox(left_frame, font=('ARIAL', 34), background='#070807', foreground='#00ff80')
    #self.Eimplante.place(x=-160, y=270, width=200, height=60)


    Euid = Entry(left_frame, font=('Arial', 24), background='#b8b8b8', justify='center' , width=1)
    Euid.place(x=-160, y=150, width=80, height=60)
   
    Euid.insert(tk.END, "7")
    EestadoAntena = Entry(left_frame, font=('ARIAL', 28), background='#ffffff', justify='center')
    EestadoAntena.place(x=-70, y=150, width=440, height=60)
    EestadoAntena.insert(0, 'Estado de la antena')
    separador = ttk.Separator(left_frame, orient=HORIZONTAL)
    separador.place(x=-160, y=225, width=530, height=3)
    self.EestadoAntena2 = Entry(left_frame, font=('ARIAL', 30), background='#070807', justify='center', foreground='#1fb1d1')
    self.EestadoAntena2.place(x=-160, y=240, width=530, height=80)
    #self.EestadoAntena2.insert(0, 'Conectada')
    login_btn = Label(left_frame, width=15, text='', font=('Times', 0), command=None, background='#003C78')
    login_btn.grid(row=10, column=1, pady=10, padx=200)

########FRAME DE LA DERECHA##############################################################################################
    right_frame = Frame(ws, bd=3,  background='#003C78', relief=SOLID, padx=320, pady=340)
    right_frame.place(x=650, y=10)
    re_pw = Label(right_frame, font=('', 14))
    re_pw.grid(row=6, column=1, pady=10, padx=20)
######pestañas############################################################################################################
    tabControl = ttk.Notebook(right_frame)
    tab1 = ttk.Frame(tabControl, width=100, height=20)
    tab5 = ttk.Frame(tabControl, width=100, height=400)
    tab6 = ttk.Frame(tabControl, width=100, height=20)
    tabControl.add(tab1, text ='Captura')
    
    tabControl.add(tab5, text ='Leidas')
    tabControl.add(tab6, text ='Registar')
    tabControl.place(x=-320, y=-340, width=687, height=590)
    #tabControl.pack(expand=1, fill="both")

    #tab5.place(x=-320, y=-340, width=687, height=590)
    Bxls = Button(right_frame, text='XLS', font=('ARIAL', 24), background='#DCE1D8', command=self.dbToXLSX, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bxls.place(x=-290, y=280, width=200, height=80)
    Bsalir = Button(right_frame, text='Salir', font=('ARIAL', 24), background='#DCE1D8', command=self.salir, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    Bsalir.place(x=90, y=280, width=200, height=80)

####tab1########################################################################################################################################################################
    self.btn1 = tk.Button(tab1, text="1",  width=6, height=2, command=lambda:self.btnClick(1),  border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn1.place(x=105, y=35)
    self.btn2 = tk.Button(tab1, text="2", width=6, height=2, command=lambda:self.btnClick(2), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn2.place(x=250, y=35)
    self.btn3 = tk.Button(tab1, text="3", width=6, height=2, command=lambda:self.btnClick(3), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn3.place(x=395, y=35)
    self.btn4 = tk.Button(tab1, text="4", width=6, height=2, command=lambda:self.btnClick(4), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn4.place(x=105, y=155)
    self.btn5 = tk.Button(tab1, text="5", width=6, height=2, command=lambda:self.btnClick(5), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn5.place(x=250, y=155)
    self.btn6 = tk.Button(tab1, text="6", width=6, height=2, command=lambda:self.btnClick(6), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn6.place(x=395, y=155)
    self.btn7 = tk.Button(tab1, text="7", width=6, height=2, command=lambda:self.btnClick(7), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn7.place(x=105, y=275)
    self.btn8 = tk.Button(tab1, text="8", width=6, height=2, command=lambda:self.btnClick(8), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn8.place(x=250, y=275)
    self.btn9 = tk.Button(tab1, text="9", width=6, height=2, command=lambda:self.btnClick(9), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn9.place(x=395, y=275)
    self.btn0 = tk.Button(tab1, text="0", width=6, height=2, command=lambda:self.btnClick(0), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btn0.place(x=250, y=395)
    self.btnborrar = tk.Button(tab1, text="C", width=6, height=2, command=self.btnClear, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btnborrar.place(x=105, y=395)
    self.btnborrar = tk.Button(tab1, text="ID", width=6, height=2, command=lambda:self.btnClick(""), border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=1, font=('ARIAL', 28), background='#DCE1D8')
    self.btnborrar.place(x=395, y=395)
    self.btnguardar = Button(tab1, text="Guardar", font=('ARIAL', 18), background='#DCE1D8', command=self.GuardarEnBD, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    self.btnguardar.place(x=560, y=35, width=100, height=120)
    self.cmb_usb=ttk.Combobox(tab1, width=10, font=('ARIAL', 28), justify='left', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:'], state='readonly')
    self.cmb_usb.place(x=560, y=300, width=100, height=60)
    self.cmb_usb.current(1)

####tab5###########################################################################################
    self.grid3 = ttk.Treeview(tab5, columns=('', '', ''), selectmode='browse')
        
    ####scrollbar##### 
    vsb = ttk.Scrollbar(tab5, orient="vertical", command=self.grid3.yview) 
    #vsb.place(x=-700, y=-100, height=500)
    vsb.pack(side='right', fill='y')
    #self.grid3 = Lis(tab5, columns=('TAG', ''))
    self.grid3.column('#0', width=178, stretch=NO)
    self.grid3.column('#1', width=200, stretch=NO)
    self.grid3.column('#2', width=150, stretch=NO)
    self.grid3.column('#3', width=150, stretch=NO)
    
    #self.grid3.place(x=-10, y=0, width=720, height=900)
    #self.grid3.place(x=-10, y=0, width=720)
    self.grid3.pack(side='left', fill='both', expand=True)
    self.grid3.heading('#0', text='Inplante', anchor='center')
    self.grid3.heading('#1', text='UHF')
    self.grid3.heading('#2', text='Peso')
    self.grid3.heading('#3', text='Estatura')
#########tab6###########################################################################################
    self.Limplante = Label(tab6, text="Implantes", font=('ARIAL', 28), background='#003C78', foreground='#ffffff', justify='center')
    self.Limplante.place(x=30, y=30)
    self.Eimplante=Entry(tab6, font=('ARIAL', 28), background='#070807', justify='right', foreground='#ffffff')
    self.Eimplante.place(x=30, y=80, width=200, height=60)
    self.Btnguardar2 = Button(tab6, text="Guardar", font=('ARIAL', 18), background='#DCE1D8', command=self.guardarImplante, border=6, highlightthickness=3, activebackground='#DCE1D8', borderwidth=6)
    self.Btnguardar2.place(x=560, y=35, width=100, height=120)


  def guardarImplante(self):
    if self.Eimplante.get() == '':
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Ingresa el implante')
    else:
      self.EestadoAntena2.delete(0, END)
      self.EestadoAntena2.insert(0, 'Ok')
      try:
        query = 'INSERT INTO implantes VALUES (?)'
        parameters = (self.Eimplante.get())
            #parameters2 = (self.grid3.item(item)['values'][1])
            #print (self.grid3.item(item)['values'][1])
        self.conexion(query, (parameters,))
        self.EestadoAntena2.delete(0, END)
        self.EestadoAntena2.insert(0, 'Implante guardado')  
        self.ListaCombo() 
            
      except:
        self.EestadoAntena2.delete(0, END)
        self.EestadoAntena2.insert(0, 'Error al guardar')

  def ListaCombo(self):
    self.lista3 = []
    with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM implantes'
      result = cursor.execute(query)
      conn.commit()
      for row in result:
        #  self.lista2.append(row[1])
         #self.Cidv['values'] = (row[1])
        
         self.lista3.append(row[0])
         #print(self.lista3)
         self.Cidv['values'] = self.lista3
         #self.grid3.insert ('', 0, text=row[0], values=(row[1], row[2], row[3]))
      cursor.close()

if __name__ == '__main__':
    ws = Tk()
    application = Principal(ws)
    ws.mainloop()