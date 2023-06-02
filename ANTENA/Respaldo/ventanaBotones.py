import binascii
import ctypes
# from distutils.cmd import Command
# from msilib.schema import Font
import os
import sqlite3
import sys
import threading
import time
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
# from tkinter import scrolledtext
# import tkinter.font as font
import tkinter.filedialog as fd
import socket
from venv import create
# from turtle import onclick
import serial
import serial.tools.list_ports
import serial.tools.list_ports_windows
import openpyxl #instalar

class Principal:
    ##nombre de la base de datos
    db_name='inventario.db'
     #se crea el socket
    # sock2 = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    # server_address2 = ('192.168.0.231', 7086)

    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server_address = ('192.168.0.231', 7086)
    sock.connect(server_address)


    # usb=os.system('fsutil fsinfo drives')
    
    # print ("cesar ",usb)
   

    def show_selection(self):
    # Obtener la opción seleccionada.
        selection = self.cmb_usb.get()
        if selection == 'C:':
            ruta = 'C:'
            
        elif selection == 'D:':
            ruta = 'D:'
        elif selection == 'E:':
            ruta = 'E:'
        elif selection == 'F:':
            ruta = 'F:'
        elif selection == 'G:':
            ruta = 'G:'
        elif selection == 'H:':      
            ruta = 'H:'
        #self.guardar_excel(ruta)
        # messagebox.showinfo(
        # message=f"La opción seleccionada es: {selection}",
        # title="Selección"
######funcion para guardar en un excel los datos leidos###############################################################################################
    def create_database(self):
        conn = sqlite3.connect(self.db_name)
        c = conn.cursor()
        c.execute("CREATE TABLE IF NOT EXISTS tags(id INTEGER PRIMARY KEY, numero TEXT)")
        conn.commit()
        conn.close()
    
    def guardar_excel(self): 
        wb = openpyxl.Workbook()    #crea el libro de excel
        sheet = wb.active  #selecciona la hoja activa
        sheet.title = 'Tags' #se le asigna un nombre a la hoja
        sheet['A1'] = 'Tags' #se le asigna un titulo a la columna A
        sheet['B1'] = 'Fecha' #se le asigna un titulo a la columna B
        sheet['C1'] = 'Hora' #se le asigna un titulo a la columna C
        
        #se le asigna un valor a cada celda
        for i in range(2, len(self.btn.get("1.0", "end-1c").splitlines())+2): 
            sheet.cell(row=i, column=1).value = self.btn.get("1.0", "end-1c").splitlines()[i-2] # i-2 para que no se repitan los datos
            sheet.cell(row=i, column=2).value = time.strftime("%d/%m/%Y")
            sheet.cell(row=i, column=3).value = time.strftime("%H:%M:%S")
            
            fecha = time.strftime("%d-%m-%Y")
            hora = time.strftime("%Hhrs%Mmin")
            extension = ".xlsx"

        selection = self.cmb_usb.get()
        if selection == '':
            messagebox.showinfo(message="Seleccione una unidad de disco", title="Selección")
       
        #wb.save(f'tags{hora}{fecha}{extension}') #se guarda el archivo en la carpeta actual
        else:
            try:
                 wb.save(selection+'\Tags-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
                 messagebox.showinfo("Excel", "Se ha guardado el archivo excel")
            except:
                messagebox.showinfo(message="No se encontro la unidad usb", title="Error")
        #print (ruta)
        #wb.save(filename='Tags.xlsx') #se guarda el libro de excel
        
        #wb.close() #se cierra el libro de excel
######funcion para guardar en un txt los datos leidos###############################################################################################
    # def guardar_contenido(self):
    #     archivo = fd.asksaveasfile(mode="w", defaultextension=".txt")
    #     if archivo is not None:
    #         archivo.write(self.txt_contenido.get(1.0, ttk.END))
    #         archivo.close()
    #     else:
    #         pass
#####se crea el array de bytes y se envian al socket para que inicie la lectura y se crea un hilo para que se ejecute en paralelo
    def IniciarLectura(self):
       
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        hilo1=threading.Thread(target=self.GetData).start()

        # lib = ctypes.CDLL('C:/Users/Mantenimiento SCR/Desktop/c#/RFIDAPI.dll')
        
        # hp = ctypes.c_void_p()
        # ip  = '192.168.0.238'
        # puerto = 7086
        # ReadType = 0
        # NoAntena = bytes(1)
        # SAAT_READ_TYPE = 1
        # SAAT_SCAN_TRIES = bytes(3)
        # lib.SAAT_TCPInit(ctypes.byref(hp),ip,puerto)
        # lib.SAAT_Open(ctypes.c_void_p())
        # lib.SAAT_6CReadTIDCode (hp, NoAntena, SAAT_READ_TYPE, SAAT_SCAN_TRIES)
        self.sock.send(Read)
        #self.sock2.send(Read)
        #read = 
#################funcion para recibir los datos del socket y mostrarlos en el textbox###############################
    def GetData(self):
        lista = []
        leidas = 5
        while True:
         data = self.sock.recv(4096)  
         #data2 = self.sock2.recv(4096)
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             #print ("Card Scanned. Tag ID:", data)
             if data not in lista:
              lista.append(data)
              #print ("guardadas",lista)
              leidas = (len(lista))
              self.txtleidas.delete(0, END)  #borra el contenido del textbox de 'leidas' para que no se repitan los datos
              self.txtleidas.insert(tk.END, leidas) #inserta el numero de datos leidos en el textbox
              #print ("leidas: ",len(lista))
              self.btn.insert(tk.END, data+"\n") #inserta los datos en el textbox
         
        #  data = str((data))
        #  lista.append(data)
        #  self.btn.insert(tk.END, data+"\n") #inserta los datos en el textbox
    def run(self):
     threading.Thread(target=self.GetData).start()

####funccion para detener la lectura##############################################################################################################################
    def DetenerLectura(self):
       
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)
#########funcion principal###############################################################################################################################
    def __init__(self, window):

        lib = ctypes.CDLL('C:/Users/Mantenimiento SCR/Desktop/c#/RFIDAPI.dll')
        
        hp = ctypes.c_void_p()
        pHostName  = '192.168.0.231'
        puerto = 7086
        nsocketPort = 7086
        ReadType = 0
        NoAntena = bytes(1)
        SAAT_READ_TYPE = 1
        SAAT_SCAN_TRIES = bytes(3)
        lib.SAAT_TCPInit(ctypes.byref(hp), pHostName, nsocketPort)
        #lib.SAAT_TCPInit(ctypes.byref(hp),ip,puerto)
        #lib.SAAT_Open(ctypes.c_void_p())
        # lib.SAAT_6CReadTIDCode (hp, NoAntena, SAAT_READ_TYPE, SAAT_SCAN_TRIES)

        self.wind = window
        self.wind.title('Agroplus')
        self.wind.geometry('1090x460')
        window.configure(background='#05183c')
       
################# funcion para que sea responsive el tamaño de la ventana ######################################################
        n_rows = 9
        n_columns = 3

        for i in range(n_rows):
            window.grid_rowconfigure(i, weight=1)
        for i in range(n_columns):
            window.grid_columnconfigure(i, weight=1, )
#####################Elementos de la ventana###################################################################################
        
        frame = LabelFrame(self.wind, text='Registrar Productos').grid(row=0, column=0, )
        #ttk.Button(frame, text='OK', ).grid(row=0, column=0, sticky=W + E, ipady=9, ipadx=120, pady=5, padx=5)
        self.btn = tk.Text(frame, width=10, height=20, font=('Arial', 32), border=3, relief=tk.GROOVE, background='#9ae5f3')
        self.btn.grid(row=0, column=0, sticky=W + E, ipady=9, ipadx=120, rowspan=6, padx=15, pady=15)
        #Button(frame, text='CONECTAR ANTENA', ).grid(row=9, column=0, sticky=W + E, ipady=9, ipadx=120, pady=5,padx=5)
        tk.Button(frame, text='SALIR', command=self.salir, width=11, font=('Arial', 24)).grid(row=5, column=2, ipady=40, ipadx=60, padx=5, pady=10)
        btn3=Button(frame, text='   LEER   ', command=self.IniciarLectura, font=('Arial', 24))
        btn3.grid(row=0, column=2, ipady=40, ipadx=60, pady=10, padx=5)
        btn3.config( height = 1, width = 11 )

        btndesconectar=tk.Button(frame, text='DEJAR DE LEER', command=self.DetenerLectura, font=('Arial', 24)).grid(row=1, column=2, ipady=40, ipadx=30, pady=1, padx=5)
        tk.Button(frame, text='    GRABAR      ', command=self.GuardarEnBD, font=('Arial', 24) ).grid(row=2, column=2, ipady=40, ipadx=43, pady=10, padx=5)
        ver=tk.Button(frame, text='          VER        ', command=self.VerDatos, font=('Arial', 24) ).grid(row=3, column=2, ipady=40, ipadx=42, pady=10, padx=5)

        tk.Button(frame, text='    RESPALDO  ', command=self.guardar_excel, font=('Arial', 24) ).grid(row=4, column=2, ipady=40, ipadx=42, pady=10, padx=5)
        Label(frame, text='LEIDAS:', bg="#05183c", fg="white", font=("Arial", 30)).grid(row=0, column=1, ipady=20, ipadx=60, padx=50, pady=20)
        self.txtleidas=tk.Entry(frame, width=10, bg="white", fg="black", font=('Arial', 16), justify='center')
        self.txtleidas.grid(row=1, column=1, ipady=9, ipadx=20)

        #txtleidas.insert(0, tleidas)
        #txtleidas.insert(self.lista
        Label(frame, text='TOTAL:', bg="#05183c", fg="white", font=("Arial", 30 )).grid(row=2, column=1, ipady=40,                                                                               ipadx=120)
        lbtotal=Entry(frame,  width=20,  bg="white", fg="white", justify='center')
        lbtotal.grid(row=3, column=1, ipady=9, ipadx=20)

        Label(frame, text='USB:', bg="#05183c", fg="white", font=("Arial", 30)).grid(row=4, column=1, ipady=40,                                                                           ipadx=120)
        #ttk.Combobox(frame, width=10, bg="white", fg="white").grid(row=5, column=1, ipady=9, ipadx=20)
        self.cmb_usb=ttk.Combobox(frame, width=10, justify='center', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:'], state='readonly')
        self.cmb_usb.grid(row=5, column=1, ipady=9, ipadx=20)
    #create_database() #crea la base de datos
        

########################Funcion para guardar los registros en la base de datos###############################################################
    def select_usb(self):
        usb = os.system('fsutil fsinfo drives')
        usb = usb.split('\n')
        listaUSB = [ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:']

        print ("cesar ",usb)
        return listaUSB

          
    
    def conexion(self, query, parameters):
     try:
      with sqlite3.connect(self.db_name) as conn:
        cursor = conn.cursor()
        result = cursor.execute(query, (parameters,))
        conn.commit()
        return result
     except Exception as e:
        messagebox.showerror("ERROR", "Error al conectar con la base de datos")
        self.create_database()
        #print("Error de conexion: ", e)
        return None

    def GuardarEnBD(self):
        if self.btn.get("1.0", "end-1c") == '' or self.btn.get("1.0", "end-1c") == 'a':
            messagebox.showerror('Error', 'NO HAY DATOS PARA GUARDAR')
            return

        for item in self.btn.get("1.0", "end-1c").splitlines():

         query = 'INSERT INTO tags VALUES (NULL, ?)'
         #parameters = (self.btn.get("1.0", "end-1c"))
         parameters = (item)

         self.conexion(query, parameters)
        #self.get_products()
        messagebox.showinfo('AGROPLUS', 'Los datos han sido guardados en la base de datos')

######FUNCION PARA VER DATOS DE LA TABLA###############################################################################################################
    def VerDatos(self):
     newWindow = Toplevel(window)
     newWindow.title("Inventario")
     newWindow.geometry("400x400")
     Label(newWindow,  text ="Inventario").pack()
     self.tree = ttk.Treeview(newWindow, height=10, columns=2)
     self.tree.pack()
     tk.Button(newWindow, text='SALIR', width=12).pack()
     #boton de salir 
      

     with sqlite3.connect(self.db_name) as conn:
      cursor = conn.cursor()
      query = 'SELECT * FROM tags'
      result = cursor.execute(query)
      conn.commit()
      for row in result:
         self.tree.insert('',0,values=row[1], text=row[0])
     cursor.close()
#######################################################################################################################################################




#######funcion para salir de la ventana###############################################################################################################
    def salir(self):
        self.DetenerLectura()
        time.sleep(0.2)
        self.wind.destroy()
        self.sock.close()
        sys.exit()
#########funcion para obtener puertos seriales###############################################################################################################

    def serial_ports(self):
        return [p.device for p in serial.tools.list_ports_windows.comports()] #lista de puertos seriales disponibles
        #return []

    def on_select( self, event):
            # get selection from event
            print("event.widget:", event.widget.get())
            # or get selection directly from combobox
            print("comboboxes: ", ttk.Combobox.get())




if __name__ == '__main__':
    window = Tk()
    application = Principal(window)
    window.mainloop()
