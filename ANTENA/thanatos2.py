from tkinter import messagebox
import openpyxl
import time
import binascii
from sqlite3 import Row
import threading
from tkinter import *
from tkinter import ttk
import tkinter as tk
import pandas as pd 

import socket

class Principal:
  try:
     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
     server_address = ('192.168.0.231', 7086)
    #server_address = ('192.168.1.200', 100)
     sock.connect(server_address)
     estado=1
  except:
     estado=0

  def __init__(self, ws):
        self.ws = ws
        self.ws.title('Agroplus')
        #self.ws.geometry('1000x600')
        self.ws.attributes("-fullscreen", True)
        self.ws.configure(background='#044275')
        self.widgets()
        s = ttk.Style()
        s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, background='#ffffff') 
        s.configure('TNotebook', background='#ffffff')
        s.configure('TFrame',    background='#ffffff')
        s.configure('TLabel',    background='#ffffff')
        s.configure('TButton',   background='#ffffff')

        s.configure("TCombobox", fieldbackground= "black", background= "black", foreground= "black", selectbackground= "black", selectforeground= "black", arrowcolor= "black", bordercolor= "black", lightcolor= "black", darkcolor= "black", troughcolor= "black", focuscolor= "black", selectcolor= "black")
        # s.configure('TNotebook.Tab', font=('Arial','20'), padding=24, fieldbackground= "orange", background='black', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        # s.configure('TNotebook', background='#003C78', borderwidth=0, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='black')
        # s.configure('TFrame', background='#003C78')
        # s.configure('TLabel', background='#003C78')
        # s.configure('TButton', background='#000000', foreground='white', borderwidth=2, relief='flat', focuscolor='black', selectcolor='black', selectbackground='black', selectforeground='white')
        
        s.configure("Treeview.Heading", font=(None, 20))
        # s.configure("Treeview.Heading", background="#993C78")
        s.configure("Treeview", font=(None, 28))
        s.configure("Treeview", rowheight=40)

        

####funcion para conectar la antena####################################################################################################################   
 
  def sockets(self,estado):
    # try:
    #     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    #     server_address = ('192.168.0.238', 7086)
    #server_address = ('192.168.1.200', 100)
       # sock.connect(server_address)
       if estado==1:
        self.EestadoAntena2.insert(0, 'Conectada')

       else:
        self.EestadoAntena2.insert(0, 'Error')
##import xls##################################################################################################################################
  def importar(self):
        self.df = pd.read_excel('C:/LISTA9.xls', dtype=str)
        self.df = self.df.fillna(0)
        self.df = self.df.astype(str)
        self.df = self.df.values.tolist()
        self.grid.delete(*self.grid.get_children())
        for i in self.df:
            self.grid.insert('', 'end', values=i)
        

####funccion para iniciar la lectura#####################################################################################     
  def IniciarLectura(self):
       
        Read = bytes([0x55, 0x0a, 0x0d, 0x04, 0x91, 0x01, 0x01, 0x00, 0x81, 0x61])
        #Read = bytes([0x0A, 0x00, 0x02, 0x28, 0x34])
        #    = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(Read)
        hilo1=threading.Thread(target=self.GetData).start()
  
  def GetData(self):
        self.lista = []
        leidas = 5
        while True:
         data = self.sock.recv(4096)  
         if data:
            if len(data) > 10:
             data = str(binascii.hexlify(data)) #convierte los datos recibidos a hexadecimal
             data = data[16:24]
             #print ("Card Scanned. Tag ID:", data)
             if data not in self.lista:
              
              self.lista.append(data)
              #print ("guardadas",lista)
              leidas = (len(self.lista))
              self.Eleidas.delete(0, END)  #borra el contenido del textbox de 'leidas' para que no se repitan los datos
              self.Eleidas.insert(tk.END, leidas) #inserta el numero de datos leidos en el textbox

              self.grid3.insert('', 'end', values=data)
              #print ("leidas: ",len(lista))
              #self.grid3.insert(tk.END, data)
              #self.grid3.insert(tk.END, data)
              self.grid3['columns'] = 'ID'
              for column in self.grid3['columns']:
                self.grid3.heading(column, text=column)

              # for item in self.lista:
              #  self.grid3.insert('', 'end', values=item)

              #self.grid3.insert('', 'END', text="L1", values=(self.lista))
              
              self.btn.insert(tk.END, data+"\n")
            for leidas in self.lista:
                
                self.Earete_entry.delete(0, END)
                self.Earete_entry.insert(tk.END, leidas)
                self.Euid2.delete(0, END)
                self.Euid2.insert(tk.END, data)
               # self.grid3.insert(tk.END, self.lista, self.lista.index(leidas+1))
                 
              #self.Earete_entry.insert(tk.END, lista.index(i+1)) #inserta los datos en el textbox
              
              #self.Earete_entry.see(tk.END) #hace scroll hasta el final del textbox
              #self.Earete_entry.delete(0, END) #borra el contenido del textbox para que no se repitan los datos
####funccion para detener la lectura##############################################################################################################################

  def DetenerLectura(self):    
        StopRead = bytes([0x55, 0x0b, 0x0d, 0x03, 0x20, 0x00, 0x01, 0x74, 0xd4])
        self.sock.send(StopRead)

        self.lista1 = []
        self.lista2 = []
        self.lista3 = []
        for line in self.grid3.get_children():
         for value in self.grid3.item(line)['values']:
          self.lista3.append(str(value))

        for line in self.grid.get_children():
         for value in self.grid.item(line)['values']:
          self.lista1.append(str(value))

        for item in self.lista3:
         if item not in self.lista1:
            self.lista2.append(item)
            self.grid2.insert('', 'end', values=(item))
        
       # self.grid3.insert(tk.END, self.lista, self.lista.index(i+1))
####funcion para salir##################################################################################################################################
  def salir(self):
        self.ws.destroy()

  def insert_data(self):

    self.Treeview.insert('', 'end', text="L1", values=(self.lista))
  
  # def treeviewtoExcel(self):
  #       wb = openpyxl.Workbook()  
  #       ws = wb.active
  #       for col in self.grid3['columns']:
  #           ws.column_dimensions[col].width = 15
  #       for row in self.grid3.get_children():
  #           for col in self.grid3['columns']:
  #               ws.cell(row=row, column=col).value = self.grid3.item(row)[col]
  #       fecha = time.strftime("%d-%m-%Y")
  #       hora = time.strftime("%Hhrs%Mmin")
  #       extension = ".xlsx"
  #       selection = self.cmb_usb.get()
  #       if selection == '':
  #           messagebox.showinfo(message="Seleccione una unidad de disco", title="Selección")
       
  #       #wb.save(f'tags{hora}{fecha}{extension}') #se guarda el archivo en la carpeta actual
  #       else:
  #           try:
  #                wb.save(selection+'\Tags-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
  #                messagebox.showinfo("Excel", "Se ha guardado el archivo excel")
  #           except:
  #               messagebox.showinfo(message="No se encontro la unidad usb", title="Error")

  #       wb.save('C:/Users/Usuario/Desktop/Reporte.xlsx')
  #       messagebox.showinfo("Excel", "Archivo exportado a Excel")

  def guardar_excel(self): 
        wb = openpyxl.Workbook()    #crea el libro de excel
        sheet = wb.active  #selecciona la hoja activa
        sheet.title = 'Tags' #se le asigna un nombre a la hoja
        sheet['A1'] = 'Tags' #se le asigna un titulo a la columna A
        sheet['B1'] = 'Fecha' #se le asigna un titulo a la columna B
        sheet['C1'] = 'Hora' #se le asigna un titulo a la columna C
        
        #se le asigna un valor a cada celda
        for i in range(2, len(self.btn.get("1.0", "end-1c").splitlines())+2): 
            sheet.cell(row=i, column=1).value = self.btn.get("1.0", "end-1c").splitlines()[i-2] # i-2 para que no se repitan los datos
            sheet.cell(row=i, column=2).value = time.strftime("%d/%m/%Y")
            sheet.cell(row=i, column=3).value = time.strftime("%H:%M:%S")
            
            fecha = time.strftime("%d-%m-%Y")
            hora = time.strftime("%Hhrs%Mmin")
            extension = ".xlsx"

        selection = self.cmb_usb.get()
        if selection == '':
            messagebox.showinfo(message="Seleccione una unidad de disco", title="Selección")
       
        #wb.save(f'tags{hora}{fecha}{extension}') #se guarda el archivo en la carpeta actual
        else:
            try:
                 wb.save(selection+'\Tags-'+fecha+'-'+hora+'.xlsx' ) #se guarda el libro de excel
                 messagebox.showinfo("Excel", "Se ha guardado el archivo excel")
            except:
                messagebox.showinfo(message="No se encontro la unidad usb", title="Error")

#####funcion para crear widgets##################################################################################################################################
  def widgets(self):
    variable = StringVar()
    gender = ('Male', 'Female', 'Other')
    variable.set(gender[0])
    #FRAME DE LA IZQUIERDA###########################################################
    left_frame = Frame(ws, bd=3, relief=SOLID, padx=210, pady=340, background='#044275', highlightbackground="white", highlightthickness=3)
    left_frame.place(x=10, y=10)
    Luid=Label(left_frame, text="UID", font=('ARIAL', 30), background='#044275', foreground='#ffffff')
    Luid.place(x=-160, y=-280)
    Larete_uhf=Label(left_frame, text="Arete UHF", font=('ARIAL', 30), background='#044275', foreground='#ffffff')
    Larete_uhf.place(x=-10, y=-280)
    self.Earete_entry = Entry(left_frame, font=('ARIAL', 24), background='#ccd7e0')
    self.Earete_entry.place(x=-10, y=-200, width=380, height=60)
    Euid = Entry(left_frame, font=('Times', 24), background='#ccd7e0')
    Euid.place(x=-160, y=-200, width=130, height=60)
    self.Euid2 = Entry(left_frame, font=('ARIAL', 24), background='#90ccfc')
    self.Euid2.place(x=-10, y=-110, width=380, height=60)
    self.Earete_uhf2 = Entry(left_frame, font=('Times', 24), background='#90ccfc')
    self.Earete_uhf2.place(x=-160, y=-110, width=130, height=60)
    Euid = Entry(left_frame, font=('Times', 24), background='#b8b8b8')
    Euid.place(x=-160, y=150, width=80, height=60)
    EestadoAntena = Entry(left_frame, font=('ARIAL', 28), background='#ffffff', justify='center')
    EestadoAntena.place(x=-70, y=150, width=440, height=60)
    EestadoAntena.insert(0, 'Estado de la antena')
    separador = ttk.Separator(left_frame, orient=HORIZONTAL)
    separador.place(x=-160, y=225, width=530, height=3)
    self.EestadoAntena2 = Entry(left_frame, font=('ARIAL', 24), background='#ffffff', justify='center')
    self.EestadoAntena2.place(x=-160, y=240, width=530, height=80)
    self.EestadoAntena2.insert(0, 'Conectada')
    login_btn = Label(left_frame, width=15, text='', font=('Times', 14), command=None, background='#044275')
    login_btn.grid(row=20, column=1, pady=10, padx=20)

    ####FRAME DE LA DERECHA#########################################################################
    right_frame = Frame(ws, bd=3,  background='#044275', relief=SOLID, padx=320, pady=340)
    right_frame.place(x=650, y=10)
    re_pw = Label(right_frame, font=('', 14))
    re_pw.grid(row=6, column=1, pady=10, padx=20)
    
    tabControl = ttk.Notebook(right_frame)
    tab1 = ttk.Frame(tabControl, width=100, height=20)
    tab2 = ttk.Frame(tabControl, width=100, height=20)
    tab3 = ttk.Frame(tabControl, width=100, height=20)
    tab4 = ttk.Frame(tabControl, width=100, height=20)
    tabControl.add(tab1, text ='Captura')
    tabControl.add(tab2, text ='Faltantes')
    tabControl.add(tab3, text ='Sin ID')
    tabControl.add(tab4, text ='leidas')
    tabControl.place(x=-320, y=-340, width=685, height=580)

    Bxls = Button(right_frame, text='XLS', font=('ARIAL', 24), background='#ffffff', command=self.guardar_excel)
    Bxls.place(x=-290, y=280, width=200, height=60)
    Bsalir = Button(right_frame, text='SALIR', font=('ARIAL', 24), background='#ffffff', command=self.salir)
    Bsalir.place(x=90, y=280, width=200, height=60)
####tab1###########################################################################################
    Larete_uhf=Label(tab1, text="Leidas", font=('ARIAL', 28), background='#ffffff')
    Larete_uhf.place(x=30, y=30)
    self.Eleidas=Entry(tab1, font=('ARIAL', 28), background='#ffffff', justify='right')
    self.Eleidas.place(x=30, y=80, width=200, height=60)
    self.Eleidas.insert(0, '0')
    Larete_uhf=Label(tab1, text="USB", font=('ARIAL', 28), background='#ffffff')
    Larete_uhf.place(x=30, y=250)
    self.cmb_usb=ttk.Combobox(tab1, width=10, font=('ARIAL', 28), justify='left', values=[ 'C:', 'D:', 'E:', 'F:', 'G:', 'H:', 'I:'], state='readonly')
    self.cmb_usb.place(x=30, y=300, width=200, height=60)

    self.bimportar = Button(tab1, text='Importar', font=('ARIAL', 24), background='#ffffff', command=self.importar)
    self.bimportar.place(x=30, y=400, width=200, height=60)
    
    Bleer=Button(tab1, text="Leer", font=('ARIAL', 28), background='#b8b8b8', command=self.IniciarLectura)
    Bleer.place(x=400, y=60, width=200, height=80)
    Bparar=Button(tab1, text="Parar", font=('ARIAL', 28), background='#b8b8b8', command=self.DetenerLectura)
    Bparar.place(x=400, y=220, width=200, height=80)
####tab2###########################################################################################
    self.grid = ttk.Treeview(tab2, columns=(1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20), show="headings", height="10")
    self.grid.place(x=10, y=0, width=650, height=500)
    self.grid.heading(1, text="ARETE")
    self.grid.heading(2, text="ID")
    
####tab3###########################################################################################

    self.grid2 = ttk.Treeview(tab3, columns=(1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20), show="headings", height="10")
    self.grid2.place(x=10, y=0, width=650, height=500)
    self.grid2.heading(1, text="ARETE")
    self.grid2.heading(2, text="ID")
    self.grid2.heading(3, text="FECHA")
    

    # for line in self.grid.get_children(): 
    #  if self.grid.get_children() == self.grid3.get_children():
    #      print("iguales")

####tab4###########################################################################################
    # self.grid3 = ttk.Treeview(tab4, columns=(1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20), show="headings", height="10") 
    self.grid3 = ttk.Treeview(tab4)
    self.grid3.place(x=-10, y=0, width=650, height=500)
    self.btn = tk.Text(tab4, width=10, height=20, font=('Arial', 32), border=3, relief=tk.GROOVE, background='#9ae5f3')
    #self.btn.grid(row=0, column=0, sticky=W + E, ipady=9, ipadx=120, rowspan=6, padx=15, pady=15)
    #self.grid3.heading(1, text="ARETE")
    #self.grid3.heading(2, text="ID")
    # self.grid3.heading(3, text="FECHA")

    # self.btn = tk.Text(tab4, width=10, height=20, font=('Arial', 32))
    # self.btn.grid(row=0, column=0, sticky=W + E, ipady=9, ipadx=120, rowspan=6, padx=15, pady=15)

if __name__ == '__main__':
    ws = Tk()
    application = Principal(ws)
    ws.mainloop()